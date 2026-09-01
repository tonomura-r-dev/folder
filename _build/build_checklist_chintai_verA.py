# -*- coding: utf-8 -*-
"""賃貸業界資料 Ver.A（入居者集客版）公表判断チェックリスト（xlsx）

  python _build/build_checklist_chintai_verA.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説、または一次ソース未確認（公表可だが「仮説」明記・裏取りが必要）
  赤 = 出典未記載の数値、または未取得の差込枠が残っている（埋まるまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DECK = ROOT / "賃貸業界_LINEOA施策提案_VerA_入居者集客.pptx"
OUT = ROOT / "賃貸業界資料_公表判断チェックリスト_VerA.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑
Y_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # 赤
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")  # 記入セル
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (枚, タイトル, 主張の種類, 判定, 根拠・出典, 残タスク)
ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "LINE公式アカウントを検討する背景", "事実", "緑",
     "LINE国内MAU 1億人突破（2025年12月末）／開封率 約55%・メール約20%（2022年6月時点）＝いずれもLINEヤフー公式",
     "提出前に公式ページで再確認（数値は変動する）"),
    (4, "市場環境｜増やせるのは「反響あたりの成約数」だけ", "事実", "黄",
     "内見2.7件・オンライン内見32.5%・家賃＋管理費は2005年以降最高＝SUUMOリサーチセンター／部屋探し期間はSUUMO記事。すべて検索スニペット経由",
     "一次ソースで再確認。反響単価・来店率・成約率を貴社実績に差し替え"),
    (5, "ユーザー理解｜電話に出ないのは冷たいからではない", "意見", "黄",
     "「心の声」6つはDYM見解（調査データではない）。ハザードマップ48.3%のみ出典つき（スニペット経由）",
     "48.3%を一次ソースで確認。心の声は仮説である旨を口頭で補足"),
    (6, "本提案のスコープ｜できること／できないこと", "意見", "緑",
     "提案スコープの宣言。数値は記載していない", "―"),
    (7, "ニーズ調査（シーズナリティ）", "事実", "緑",
     "Googleトレンド実測（日本／直近1年・2026年8月取得）。相対指標である旨を脚注に明記",
     "提案直前に再取得して取得日を更新する"),
    (8, "ニーズ調査（前後検索）", "事実", "緑",
     "LINEヤフー 前後検索データ（Journey・直近1年／2026年8月受領）。起点KW・受領日を明記済",
     "―"),
    (9, "他社分析①｜競合のLINE運用ステータス", "事実", "赤",
     "アカウントIDは2026年8月に特定済。★友だち数が未取得で破線の差込枠のまま（クラウドから page.line.me がブロック）",
     "page.line.meで取得日つき実測（PC作業）。埋まるまで本スライドは対外提出しない"),
    (10, "他社分析②｜運用は4型。大手は店舗ごとに分散", "意見", "黄",
     "4分類はDYM見解（仮説バッジ表示済）。店舗別アカウントの存在は調査で確認済（ID特定まで）",
     "友だち数の実測後に各社をプロットして確定。競合の実名掲載可否を社内確認"),
    (11, "全体設計｜カスタマージャーニー8フェーズ × CV3段", "意見", "緑",
     "CV3段は本提案での定義（DYM設計）", "貴社のCV定義とすり合わせ"),
    (12, "施策全体像｜触るのは2箇所だけ", "意見", "緑",
     "スコープの図解。数値は記載していない", "―"),
    (13, "LINE公式アカウントにおける対策領域", "意見", "緑",
     "機能マップ（FMT）。賃貸の文脈に置換済", "―"),
    (14, "施策展開図（初期・月次）", "意見", "緑",
     "実施メニュー（FMT）。「※有料」項目はP30と対応", "―"),
    (15, "構築｜友だち追加動線", "意見", "緑",
     "5経路の動線設計と実文面。数値は記載していない",
     "LINE広告CPFの審査要件・クリエイティブ表現の制限を確認"),
    (16, "構築｜あいさつメッセージ", "意見", "黄",
     "文面はDYM設計。脚注のブロック率20〜30%は第三者メディア（本文では目標値にしていない）",
     "20〜30%を一次ソースで確認、取れなければ脚注ごと削除"),
    (17, "構築｜希望条件アンケート", "意見", "緑",
     "設問設計。1問目をエリアにする根拠はLINEヤフー前後検索データ（P8）",
     "貴社の物件データ項目に合わせて選択肢を調整"),
    (18, "構築｜リッチメニュー 3タブ×6枠", "意見", "緑",
     "18ボタンの設計と実装仕様（2500×1686px）。数値の主張なし", "―"),
    (19, "構築｜反響即時対応", "意見", "緑",
     "「5分以内」は設計値であり実績値として置いていない。分数の断定は本文にない",
     "現在の初回返信までの平均時間を実測して置き換える"),
    (20, "配信設計｜14日ステップ", "意見", "黄",
     "各日の根拠はLINEヤフー前後検索データとSUUMO。検討期間「約3週間〜1ヶ月」はSUUMO記事（スニペット経由）",
     "検討期間を一次ソースで確認"),
    (21, "配信設計｜実文面①（Day0・Day3）", "意見", "緑",
     "配信文面の初稿。物件名・家賃・初期費用は記入例である旨を脚注に明記",
     "物件データベースとの自動差し込み方法を確認"),
    (22, "配信設計｜実文面②（Day7・Day14）", "意見", "緑",
     "同上。Day14を審査にする根拠はLINEヤフー前後検索データ（P8）", "―"),
    (23, "配信設計｜物件レコメンド", "意見", "緑",
     "配信軸の設計。「初期費用が安い」を軸にする根拠はLINEヤフー前後検索データ（P8）",
     "配信頻度は運用開始後のブロック率・反応率で調整"),
    (24, "配信設計｜年間の企画投稿カレンダー", "事実", "黄",
     "参照データはGoogleトレンド実測とLINEヤフー前後検索データ。9月のオンライン内見率のみスニペット経由",
     "オンライン内見率を一次ソースで確認"),
    (25, "配信設計｜LINE通知メッセージ", "意見", "黄",
     "利用シーンの設計。他業種の公開実績（約30%・44%）は参考として脚注に置き、目標値にしていない",
     "電話番号の取得同意状況を確認。脚注の他業種実績を一次ソースで確認 or 削除"),
    (26, "歩留まり｜来店率改善", "事実", "黄",
     "オンライン内見実施率32.5%＝SUUMOリサーチセンター（2022年度・首都圏／スニペット経由）",
     "一次ソースで確認。ドタキャン率・来店率を貴社実績に差し替え"),
    (27, "歩留まり｜工数削減", "意見", "緑",
     "Before/Afterの業務整理。削減時間の想定値は置いていない",
     "架電件数・所要時間の実測をもらって「営業◯名の◯日分」まで定量化"),
    (28, "改善モデル", "意見", "緑",
     "機能設計（FMT）。★FMT由来の出典未記載数値（CVR 2%→5〜10%）は本版で削除済", "―"),
    (29, "成果｜効果測定の設計", "意見", "緑",
     "測る段と主KPI（CPO）と構造のみ。★業界汎用のため数値は一切入れていない",
     "貴社実績が入り次第、この構造に数字を入れて別紙SIMを作成"),
    (30, "体制｜費用プラン", "事実", "黄",
     "運用プラン・ツール費はDYM標準（2026年8月時点）。LINEヤフー社への支払分と分けて記載",
     "最新の料金表と一致しているか社内確認（上司確認②）"),
    (31, "体制｜運用スケジュール", "意見", "緑",
     "標準的な6ヶ月の進行", "既存アカウントの有無・API連携の要否で前後を調整"),
    (32, "体制｜サポート体制", "意見", "緑",
     "DYM標準の体制と分担", "有人チャットの対応時間・担当を運用開始前に取り決め"),
    (33, "締め｜飛び道具（4案）", "意見", "黄",
     "A〜CはDYM見解（効果を保証する裏付けなし）。Dのみ層の存在をLINEヤフー前後検索データで確認済",
     "A〜Cが仮説である旨は脚注に記載済。Dは翻訳ツールと対応体制の検討が別途必要"),
    (34, "締め｜第2の提案軸（オーナー向け）", "意見", "緑",
     "Ver.Bとの比較整理。数値は記載していない", "別冊Ver.B（オーナー開拓版）の完成が前提"),
    (35, "締め｜LINEOA実績", "事実", "黄",
     "LINEヤフー公式が実名公開している導入事例4件。数値はスニペット経由で取得",
     "掲載前に各事例ページ原典で数値を再確認（実名掲載自体は出典明記で可）"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（FMTのまま）", "―"),
]

wb = Workbook()

# ---------------- サマリ ----------------
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "賃貸業界_LINEOA施策提案 Ver.A｜入居者集客版（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-28　対象ファイル：賃貸業界_LINEOA施策提案_VerA_入居者集客.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("ポータル反響は、いま使い捨てになっている。反響数も出稿費も増やさずに、"
            "反響を資産化するだけで成約1件あたりのコスト（CPO）は下がる。"
            "触るのは「反響が来た後」と「予約が入った後」の2箇所だけ。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（仮説明記・裏取りのうえ公表可）", "黄", Y_FILL),
    ("赤（差込が埋まるまで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = ('=IF(D11=0,"OK：赤0件。対外提出可",'
             '"NG：赤"&D11&"件。差込枠が埋まるまで対外提出不可（該当スライドを抜けば提出可）")')
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 注文住宅版との違い（FMT由来の赤4数値は本版では発生していません）"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B18"] = ("FMTの出典未記載4数値（業態平均で最大3%／CVR 2%→5〜10%／商談化70%改善・再CV率3〜8%／"
             "友だち追加率0.8〜1.0%）が載っていたFMTの4枚は、本版では別内容に差し替えたか、"
             "文言から該当数値を削除しています。")
ws["B18"].font = Font(name=FONT, size=9.5)
ws["B18"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B18:F18")
ws.row_dimensions[18].height = 30

ws["B20"] = "★ 上司・社内で確認が必要な3点"
ws["B20"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① 競合の実名掲載の可否（P9・P10）→ NGなら「A社／B社」表記に変更する。"
    "※LINEヤフーが実名公開している導入事例（P35）は出典明記で掲載可なので、確認は競合分析だけでよい",
    "② 費用プラン（P30）の金額が最新のDYM料金表と一致しているか。ツール費・通知メッセージの通数単価も含む",
    "③ 脚注に残した第三者メディア由来の数値を残すか削除するか（P16のブロック率20〜30%／P25の他業種の通知メッセ実績）",
]):
    r = 21 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

ws["B25"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B25"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. 競合の友だち数の実測（IDは特定済み／取得日を併記／管理を併営しているかも記録）→ P9が赤→緑、P10が黄→緑",
    "2. 一次ソースの裏取り（SUUMOリサーチセンター 賃貸契約者動向調査／日管協短観／LINEヤフー導入事例ページ）"
    "→ P4・P20・P24・P26・P35が黄→緑",
    "3. 貴社の実績値（反響数・ポータル出稿費・来店率・成約率・ドタキャン率・架電件数と所要時間）"
    "→ P4・P26・P27が実測になり、P29の構造に数字を入れて別紙SIMが作れる",
    "4. Googleトレンドの再取得（提案直前に取得日を更新。KWは必ず一語で）→ P7・P24",
    "5. 物件データベースとの連携方法（新着物件をLINE配信へ自動差し込みできるか）→ P21・P23",
]):
    r = 26 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 26

ws["B32"] = "凡例（このファイルの使い方）"
ws["B32"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
    "・本資料は業界汎用のため、成果シミュレーション（SIM）は入れていません（P29は構造のみ）。"
    "個社提案に落とすときに、上記3の実績値をもらってSIMを作ります。",
]):
    ws.cell(33 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

# ---------------- 判定一覧 ----------------
ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 38

for col, w in zip("ABCDEF", [5, 42, 11, 7, 58, 48]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
