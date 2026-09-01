# -*- coding: utf-8 -*-
"""「既存アタック用_案件一覧.xlsx」のエンド別サマリに契約形態の列を足す。

  python _build/add_keiyaku_column.py <入力.xlsx> [出力.xlsx]

佐村さんからこの形式でOKが出たので、足りなかった契約形態だけを追加する。

足す列
  契約形態     : ストック / ショット / ストック＋ショット
  ストック売上 : ストック案件ぶんの売上
  ショット売上 : ショット案件ぶんの売上

数字は SUMIFS の数式でも書けるが、795行 × 1,516件を集計すると
Excelで開くたびに固まるため、元ファイルと同じく「値」で書き込む。
最新化したいときはこのスクリプトを流し直す。
"""
import sys
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

FONT = "メイリオ"
NAVY = "1F3864"
YEN = "#,##0;-#,##0;\"-\""
PCT = "0.0%"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SUM_SHEET = "エンド別サマリ"
DET_SHEET = "案件一覧"
HEAD_ROW = 3            # 両シートとも見出しは3行目、データは4行目から
FIRST = HEAD_ROW + 1
KEIS = ["ストック", "ショット"]

HEADERS = ["No.", "エンドクライアント名", "契約形態", "案件数", "売上",
           "ストック売上", "ショット売上", "粗利", "粗利率",
           "媒体", "商材カテゴリ", "担当"]
WIDTHS = [6, 34, 16, 8, 14, 14, 14, 13, 9, 22, 26, 30]
MONEY = (5, 6, 7, 8)
WRAP = (10, 11, 12)


def f(sz=9, bold=False, color="000000"):
    return Font(name=FONT, size=sz, bold=bold, color=color)


def norm(v):
    return "" if v is None else str(v).strip()


def num(v):
    return v if isinstance(v, (int, float)) else 0


def head_map(ws):
    """見出し名 → 1始まりの列番号。列が動いていても壊れないように名前で引く。"""
    row = next(ws.iter_rows(min_row=HEAD_ROW, max_row=HEAD_ROW,
                            max_col=40, values_only=True))
    out = {}
    for i, v in enumerate(row, start=1):
        name = norm(v).replace("\n", "")
        if name and name not in out:
            out[name] = i
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "既存アタック用_案件一覧_契約形態あり.xlsx"

    wb = load_workbook(src)
    det, smy = wb[DET_SHEET], wb[SUM_SHEET]

    d = head_map(det)
    need = ["エンドクライアント名", "契約形態", "売上", "粗利",
            "媒体", "商材カテゴリ"]
    missing = [k for k in need if k not in d]
    if missing:
        raise SystemExit(f"案件一覧に見つからない列: {missing}")
    tanto_col = next((i for n, i in d.items() if n.startswith("担当")), None)

    # 案件一覧を1回だけ舐めて、エンドごとに集計する
    g = defaultdict(lambda: {"n": 0, "uri": 0.0, "rieki": 0.0,
                             "ストック": 0.0, "ショット": 0.0, "kei": set(),
                             "媒体": [], "商材": [], "担当": []})
    for row in det.iter_rows(min_row=FIRST, max_col=det.max_column, values_only=True):
        end = norm(row[d["エンドクライアント名"] - 1])
        if not end:
            continue
        g[end]["n"] += 1
        g[end]["uri"] += num(row[d["売上"] - 1])
        g[end]["rieki"] += num(row[d["粗利"] - 1])
        kei = norm(row[d["契約形態"] - 1])
        if kei in KEIS:
            # 契約形態は金額ではなく種別そのもので判定する
            # （金額0のヨミ案件を取りこぼさないため）
            g[end][kei] += num(row[d["売上"] - 1])
            g[end]["kei"].add(kei)
        for src_col, dst in (("媒体", "媒体"), ("商材カテゴリ", "商材")):
            v = norm(row[d[src_col] - 1])
            if v and v not in g[end][dst]:
                g[end][dst].append(v)
        if tanto_col:
            for t in norm(row[tanto_col - 1]).split(" / "):
                if t and t not in g[end]["担当"]:
                    g[end]["担当"].append(t)

    def label(kei):
        if len(kei) == 2:
            return "ストック＋ショット"
        if "ストック" in kei:
            return "ストック"
        if "ショット" in kei:
            return "ショット"
        return "（未設定）"

    order = sorted(g, key=lambda k: -g[k]["uri"])

    # サマリシートを作り直す（元から自動生成の表なので手入力は無い）
    smy.delete_rows(HEAD_ROW, smy.max_row - HEAD_ROW + 2)
    if smy.auto_filter:
        smy.auto_filter.ref = None

    for i, h in enumerate(HEADERS, start=1):
        c = smy.cell(row=HEAD_ROW, column=i, value=h)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    smy.row_dimensions[HEAD_ROW].height = 30

    for j, name in enumerate(order):
        r, v = FIRST + j, g[name]
        vals = [j + 1, name, label(v["kei"]), v["n"], v["uri"],
                v["ストック"], v["ショット"], v["rieki"],
                (v["rieki"] / v["uri"] if v["uri"] else None),
                " / ".join(v["媒体"]), " / ".join(v["商材"]), " / ".join(v["担当"])]
        for i, val in enumerate(vals, start=1):
            c = smy.cell(row=r, column=i, value=val)
            c.font = f(9)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in WRAP))
        for i in MONEY:
            smy.cell(row=r, column=i).number_format = YEN
        smy.cell(row=r, column=9).number_format = PCT
        smy.cell(row=r, column=3).alignment = Alignment(horizontal="center",
                                                        vertical="top")

    dual = sum(1 for k in order if len(g[k]["kei"]) == 2)
    smy["A1"] = "エンド別サマリ｜既存アタックの優先順位（売上の大きい順）"
    smy["A1"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
    smy["A2"] = (f"{len(order)}社（うち{dual}社はストックとショットの両方を保有）"
                 "／金額は税抜／契約形態は案件一覧から判定")
    smy["A2"].font = f(9, color="44546A")

    for i, w in enumerate(WIDTHS, start=1):
        smy.column_dimensions[get_column_letter(i)].width = w
    smy.freeze_panes = f"C{FIRST}"
    smy.auto_filter.ref = (f"A{HEAD_ROW}:"
                           f"{get_column_letter(len(HEADERS))}{HEAD_ROW + len(order)}")

    wb.save(out)

    from collections import Counter
    c = Counter(label(g[k]["kei"]) for k in order)
    print(f"書き出しました: {out}")
    print(f"  エンド社数    : {len(order):,}")
    print(f"  契約形態の内訳: {dict(c)}")
    print(f"  売上合計      : {sum(g[k]['uri'] for k in order):,.0f}")
    print(f"  ストック計    : {sum(g[k]['ストック'] for k in order):,.0f}")
    print(f"  ショット計    : {sum(g[k]['ショット'] for k in order):,.0f}")


if __name__ == "__main__":
    main()
