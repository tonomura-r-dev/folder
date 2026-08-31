# -*- coding: utf-8 -*-
"""月次管理ブックの既存顧客を「LINEOA提案管理シート」の行の形に組み替える。

  python _build/build_teian_import.py <月次管理ブック.xlsx> [出力先ベース名]
  オプション:
    --月 26/8                提案列に入れる年月（既定は今月）
    --商材 商流変更          提案商材に入れる値
    --除外アカウント貸し     サービス詳細が「アカウント貸し」の案件を外す

■ シートの列の意味（実データから読み取ったもの。ここを間違えると全部ズレる）
    提案     提案した年月。`23/12` `26/7` の形
    結論     日付だったり `NG` `未提案` `ー` だったりで統一されていない → 空欄
    ☑        カウント（同じ会社が何回出ているか）。新しい行は 1
    追客     ほぼ空欄
    企業名   会社名
    Ac名     LINE公式アカウントの名前。実績データに無いので空欄
    既存     `●` `△`。既存顧客の印
    商流     いま経由している代理店の「会社名そのもの」。直なら `直`、不明なら `不明`
    外注先   DYM側のパートナー（エフ・コード、LINY、第一通信社など）。データに無いので空欄
    提案商材 `商流変更` `コンサル` `CPF` `成果報酬` `代理店` `既存伸長（AD）`
    業界     客の業種（飲食・アパレル・医療…）。実績データに無いので空欄
    担当①②  **名字だけ**。フルネームでは入れない
    DYM売上  提案時の見込み額。まだ提案していない行は空欄
    LINE売上 同上

■ 埋める列と、空欄のままにする列
    埋める : 提案 / ☑ / 企業名 / 既存 / 商流 / 提案商材 / 担当① / 担当②
    空欄   : 結論 / 追客 / Ac名 / 外注先 / 業界 / DYM売上 / LINE売上
             実績データから出せない、または「これから決めること」なので入れない。
             業界だけはシート側に既に入っているので、貼ったあと VLOOKUP で引ける。

出力
  <ベース名>.xlsx           15列＋参考列。貼る前の確認用
  <ベース名>_貼り付け用.tsv  15列ちょうど・見出し無し。そのまま行に貼れる
  <ベース名>_参考列.tsv      参考列だけ。P列以降に貼りたいとき用

元ブックは読むだけで一切変更しない。
"""
import sys
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_attack_list import extract  # 抽出処理は既存アタック用リストと共通

FONT = "メイリオ"
NAVY = "1F3864"
GRAY = "808080"
LIGHT = "F2F2F2"
YEN = "#,##0;-#,##0;\"-\""
PCT = "0.0%"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 提案管理シートの見出し。並びを変えると貼り付け位置がずれるので触らない
SHEET_COLS = ["提案", "結論", "☑", "追客", "企業名", "Ac名", "既存", "商流",
              "外注先", "提案商材", "業界", "担当①", "担当②", "DYM売上", "LINE売上"]
# 判断材料として右に足す列。貼るかどうかは任意
REF_COLS = ["契約形態", "実績売上", "実績粗利", "粗利率", "案件数", "媒体",
            "商材カテゴリ", "担当（フルネーム）", "最新対象月", "アカウント貸し"]

WIDTHS = ([9, 9, 5, 8, 30, 14, 6, 24, 14, 14, 10, 10, 10, 12, 12] +
          [16, 14, 13, 9, 8, 20, 24, 26, 12, 12])

LINE_MEDIA = "LINE公式アカウント"
KAIGASHI = "アカウント貸し"
KEIS = ["ストック", "ショット"]

# 担当の名字。提案管理シートの担当①②に実際に入っている表記をそのまま使う。
# フルネーム（奥田大輝）を頭から最長一致で削って名字（奥田）にするための辞書。
# 3文字の名字を先に見るので、長いものから並べる必要はない（下でソートする）。
SURNAMES = """
長谷川 白拍子 阿久津 百目鬼 小野里 オルネラス 水小田 佐々木
奥田 大友 田畑 小澤 池田 伊藤 薗田 星川 本山 佐村 首藤 茨城 平田 富田
武舍 武舎 小池 向井 石井 五味 半田 櫻井 木村 杉山 矢野 山口 疋田 四方
榊原 飯川 入澤 村形 渡邊 渡辺 尾勝 鵜殿 古賀 森下 齋藤 齊藤 寺島 清水
金子 小野 小林 中山 高橋 髙橋 福井 加藤 大西 森川 納富 納冨 槌谷 青木
山田 床波 石田 垣内 松岡 鈴木 田中 海貝 最上 大貫 吉田 北川 竹田 矢代
陶山 桑野 早崎 中村 櫻田 紙野 川本 平井 小川 佐藤 山内 藤浪 高島 太田
飯島 松本 伊賀 岩崎 細野 本宮 宮本 宮元 武智 星澤 木下 菅川 高田 高
渡邊祥 松元 宮澤 星 林 宋 東 金 柳 塩谷 石田 疋田 小田 三宅 大石
""".split()


def f(sz=10, bold=False, color="000000"):
    return Font(name=FONT, size=sz, bold=bold, color=color)


def surname(full):
    """フルネームから名字だけを取り出す。シートの担当欄が名字表記のため。

    「伊藤(優)」のように括弧つきで区別している人がいるので、元の表記に
    括弧があればそれごと残す。辞書に無い名前は先頭2文字で妥協する。
    """
    s = str(full or "").strip()
    if not s:
        return ""
    if "(" in s or "（" in s:   # 伊藤(優) などは既に名字表記
        return s
    for n in sorted(SURNAMES, key=len, reverse=True):
        if s.startswith(n):
            return n
    return s[:2] if len(s) > 2 else s


def month_key(s):
    """「8月分」などから並べ替え用の数値を作る。判別できないものは0。"""
    digits = ""
    for ch in s:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    return int(digits) if digits else 0


def push(lst, v):
    if v and v not in lst:
        lst.append(v)


def aggregate(recs, teian, shozai):
    """エンドクライアント単位で1行にまとめる。売上の大きい順＝アタックの優先順。"""
    g = defaultdict(lambda: {
        "売上": 0.0, "LINE売上": 0.0, "粗利": 0.0, "件数": 0, "貸し": 0,
        "直": False, "代理店名": [], "kei": set(),
        "媒体": [], "商材": [], "担当": [], "月": [], "acct": "", "cons": "",
    })

    for r in recs:
        end = r["エンドクライアント名"]
        if not end:
            continue
        d = g[end]
        d["件数"] += 1
        d["売上"] += r["売上"]
        d["粗利"] += r["粗利"]
        if r["媒体"] == LINE_MEDIA:
            d["LINE売上"] += r["売上"]
        if KAIGASHI in r["サービス詳細"]:
            d["貸し"] += 1

        # 商流はシートでは「経由している代理店の会社名そのもの」。直なら「直」
        if "直" in r["種別"]:
            d["直"] = True
        elif "代理店" in r["種別"] and r["社名"] and r["社名"] != end:
            push(d["代理店名"], r["社名"])

        if r["契約形態"] in KEIS:
            d["kei"].add(r["契約形態"])
        push(d["媒体"], r["媒体"])
        push(d["商材"], r["商材カテゴリ"])
        push(d["月"], r["対象月"])
        for t in r["担当"].split(" / "):
            push(d["担当"], t)
        if not d["acct"]:
            d["acct"] = r["アカウント担当"]
        if not d["cons"]:
            d["cons"] = r["コンサル担当"]

    def keiyaku(s):
        if len(s) == 2:
            return "ストック＋ショット"
        return next(iter(s)) if s else "（未設定）"

    rows = []
    for name in sorted(g, key=lambda k: -g[k]["売上"]):
        d = g[name]
        # 直の実績があれば「直」。無ければ代理店名を並べる。どちらも無ければ不明
        shoryu = "直" if d["直"] else (" / ".join(d["代理店名"]) or "不明")
        sheet = [teian, "", 1, "",                    # 提案 / 結論 / ☑ / 追客
                 name, "", "●", shoryu,               # 企業名 / Ac名 / 既存 / 商流
                 "", shozai, "",                      # 外注先 / 提案商材 / 業界
                 surname(d["acct"]), surname(d["cons"]),   # 担当① / 担当②
                 "", ""]                              # DYM売上 / LINE売上（未提案）
        ref = [keiyaku(d["kei"]), d["売上"], d["粗利"],
               (d["粗利"] / d["売上"] if d["売上"] else None), d["件数"],
               " / ".join(d["媒体"]), " / ".join(d["商材"]),
               " / ".join(d["担当"]),
               max(d["月"], key=month_key) if d["月"] else "",
               f"●（{d['貸し']}件）" if d["貸し"] else ""]
        rows.append(sheet + ref)
    return rows


# ============================================================ 書き出し
def write_xlsx(path, rows, teian, shozai, excluded):
    wb = Workbook()
    ws = wb.active
    ws.title = "提案管理シート用"

    ws["A1"] = "提案管理シート用｜列の並びと書き方を提案管理シートに合わせたもの"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = (f"{len(rows)}社／A〜O列が提案管理シートと同じ並び。P列から右は判断材料で、"
                f"貼るかどうかは任意。提案={teian}、提案商材={shozai}、担当は名字表記。"
                "DYM売上・LINE売上は提案時の見込み額なので未提案のうちは空欄。"
                + ("／アカウント貸しの案件は除外済み" if excluded else ""))
    ws["A2"].font = f(9, color="44546A")

    head = SHEET_COLS + REF_COLS
    for i, h in enumerate(head, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY if i <= len(SHEET_COLS) else GRAY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 30

    money = (14, 15, 17, 18)
    wrap = (8, 21, 22, 23)
    for j, row in enumerate(rows):
        rr = 4 + j
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=rr, column=i, value=v)
            c.font = f(9)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in wrap))
            if i > len(SHEET_COLS):
                c.fill = PatternFill("solid", fgColor=LIGHT)
        for i in money:
            ws.cell(row=rr, column=i).number_format = YEN
        ws.cell(row=rr, column=19).number_format = PCT
        for i in (3, 7):
            ws.cell(row=rr, column=i).alignment = Alignment(horizontal="center")

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "F4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(head))}{3 + len(rows)}"
    wb.save(path)


def cell(v, pct=False):
    if v is None:
        return ""
    if pct:
        return f"{v:.4f}"
    if isinstance(v, float):
        return f"{v:.0f}"
    return str(v)


def write_tsv(path, rows, lo, hi, pct_at=None):
    """rows の [lo:hi] だけを書き出す。見出しは付けない（そのまま行に貼るため）。"""
    with open(path, "w", encoding="utf-8") as fh:
        for row in rows:
            fh.write("\t".join(cell(v, pct=(i == pct_at))
                               for i, v in enumerate(row[lo:hi], start=lo)) + "\n")


def opt(name, default):
    """--名前 値 の形のオプションを読む。"""
    if name in sys.argv:
        i = sys.argv.index(name)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return default


def main():
    today = date.today()
    teian = opt("--月", f"{today.year % 100}/{today.month}")
    shozai = opt("--商材", "商流変更")
    excluded = "--除外アカウント貸し" in sys.argv

    skip = {teian, shozai, "--月", "--商材"}
    args = [a for a in sys.argv[1:] if not a.startswith("--") and a not in skip]
    if not args:
        print(__doc__)
        sys.exit(1)
    src = args[0]
    base = args[1] if len(args) > 1 else "提案管理シート_移管用"

    recs, skipped = extract(src)
    if excluded:
        before = len(recs)
        recs = [r for r in recs if KAIGASHI not in r["サービス詳細"]]
        print(f"  アカウント貸しを除外: {before - len(recs)}件")

    rows = aggregate(recs, teian, shozai)
    n = len(SHEET_COLS)
    write_xlsx(f"{base}.xlsx", rows, teian, shozai, excluded)
    write_tsv(f"{base}_貼り付け用.tsv", rows, 0, n)
    write_tsv(f"{base}_参考列.tsv", rows, n, len(rows[0]), pct_at=n + 3)

    choku = sum(1 for r in rows if r[7] == "直")
    fumei = sum(1 for r in rows if r[7] == "不明")
    kashi = sum(1 for r in rows if r[n + 9])
    print(f"書き出しました: {base}.xlsx / {base}_貼り付け用.tsv / {base}_参考列.tsv")
    print(f"  社数              : {len(rows):,}")
    print(f"  提案（年月）      : {teian}")
    print(f"  提案商材          : {shozai}")
    print(f"  商流              : 直 {choku}社 ／ 代理店名あり {len(rows)-choku-fumei}社 "
          f"／ 不明 {fumei}社")
    print(f"  担当①が空欄      : {sum(1 for r in rows if not r[11])}社")
    print(f"  アカウント貸しあり: {kashi}社")
    print(f"  実績売上 合計     : {sum(r[n+1] for r in rows):,.0f}（参考列。本体は空欄）")
    for sheet, miss in skipped:
        print(f"  ★スキップ {sheet}: {miss}")


if __name__ == "__main__":
    main()
