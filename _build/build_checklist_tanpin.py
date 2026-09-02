# -*- coding: utf-8 -*-
"""単品リピート通販業界資料 公表判断チェックリスト（xlsx）

  python3 _build/build_checklist_tanpin.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説。データが入れば緑になる（公表可だが「仮説」「未取得」明記が必要）
  赤 = 出典未記載の数値が載っている（出典を確定 or 削除するまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DECK = ROOT / "単品通販業界_LINEOA施策提案.pptx"
OUT = ROOT / "単品通販業界資料_公表判断チェックリスト.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑
Y_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # 赤
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")  # 記入セル
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (枚, タイトル, 主張の種類, 判定, 根拠・出典, 残タスク)
ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "市場データ｜ニーズ全体図", "事実", "緑", "日本ネット経済新聞／日本流通産業新聞「2025年版 通販売上高ランキング」（化粧品通販は2024年4月〜2025年3月決算・91社）", "提案時に最新年版を確認"),
    (4, "単品通販の構造｜3つの壁", "事実", "緑", "初回離脱率・F2転換率＝単品通販のCRM各社／相談件数＝国民生活センター／CPC・広告費＝キーワードマーケティング・電通", "相談件数の最新年度を確認（kokusen.go.jp）"),
    (5, "顧客心理｜なぜ1回で終わるのか", "事実", "緑", "継続率の商材差＝定期通販CRM各社／相談件数＝国民生活センター／メルマガ開封率＝国内EC平均／開封率はLINEヤフー公式値", "―"),
    (6, "できること／できないこと＋スコープ宣言", "意見", "緑", "DYM提案のスコープ宣言。「解約は止めない」を明記。数値なし", "―"),
    (7, "実態調査①｜広告単価は、もう下がらない", "事実", "緑", "電通「2025年 日本の広告費」2026年3月5日発表／キーワードマーケティング。限界CPOの定義は業界標準用語", "―"),
    (8, "実態調査②｜弊社の広告実績（CV地点別）", "推定", "黄", "★差込枠。数値は未記入（捏造なし）", "DYM社内の単品通販広告実績をCV地点別（初回購入／定期引上／F2）に取得（上司確認①）"),
    (9, "実態調査③｜法規制と広告審査", "事実", "緑", "特定商取引法改正（2021年成立・2022年6月1日施行）／薬機法の広告3要件（薬監発第148号）／景品表示法", "薬機法・特商法の審査落ち実例（社内1〜2例）があれば説得力が上がる（上司確認②）"),
    (10, "実態調査④｜検索の季節性", "推定", "黄", "★差込枠。KWのみ提示、データ未記入", "Googleトレンド実データ（サプリ／青汁／定期便／解約・一語・5年＋1年。スケール別に分ける）"),
    (11, "実態調査⑤｜前後検索", "推定", "黄", "★差込枠。KWのみ提示、データ未記入", "LINEヤフー前後検索データ（サプリ／定期便／解約／口コミ。★「解約」が本命）"),
    (12, "実態調査⑥｜競合のLINE運用", "事実", "黄", "運用の型4分類はDYM整理（意見）。競合8社の選定は知名度優先（事実）。友だち数・解約導線は★未取得", "page.line.meから取得日つきで実測（友だち数／リッチメニューに解約があるか／解約がLINEで完結するか）"),
    (13, "全体設計①｜フェーズ8分類×CV3段", "意見", "緑", "CV3段はDYM設計。数値なし", "貴社のCV定義とすり合わせ"),
    (14, "全体設計②｜施策全体像と対策領域", "意見", "緑", "DYM提案の対策領域マップ。施策とKPIの対応表", "―"),
    (15, "全体設計③｜施策展開図（初期・月次）", "意見", "緑", "DYM標準の展開図を単品通販施策に置換", "―"),
    (16, "構築①｜友だち追加動線（同梱物QRが主役）", "意見", "緑", "DYM提案の動線設計", "―"),
    (17, "構築②｜あいさつメッセージ", "意見", "緑", "配信文面は初稿。「いつでも解約可能」等の強調表示を使わない旨は特商法に基づく", "担当者名・実際の発送予定に差し替え"),
    (18, "構築③｜初期アンケートとタグ設計", "意見", "緑", "設問設計。数値なし。F2の逆算方法のみ提示", "商材ごとに内容量÷1日使用量で日数を確定"),
    (19, "構築④｜リッチメニュー", "意見", "緑", "DYM提案のリッチメニュー設計。「解約・休止」を1軍に置く方針", "―"),
    (20, "構築⑤｜LINE ID連携（★前提）", "意見", "緑", "仕組みの説明。数値なし", "カート／定期システム側の連携可否を初期に確認"),
    (21, "配信設計①｜シナリオ3本の設計表", "意見", "緑", "配信設計。日数は書かず逆算の型のみ（★想定を並べていない）", "F2オファーのタイミングは商材ごとに調整"),
    (22, "配信設計②｜実文面（初回購入後→F2）", "意見", "緑", "配信文面は初稿。効果効能を断定しない書き方にしている", "★薬機法チェックを通すこと。実際の商品名・日数に差し替え"),
    (23, "配信設計③｜実文面（定期継続中・解約予防）", "意見", "緑", "配信文面は初稿。スキップ提示は解約妨害に当たらない設計", "実際の発送サイクルに差し替え"),
    (24, "配信設計④｜解約シナリオ（★核心）", "意見", "緑", "特商法（契約解除の妨害に罰則）に基づく設計。チャーン削減幅1〜2%は業界水準", "★「止めない」の明記を消さないこと。法務確認を推奨"),
    (25, "配信設計⑤｜年間の企画投稿カレンダー", "推定", "黄", "現時点の企画は根拠列が未確定の叩き台", "S10のGoogleトレンド確定後に「参照データ」列を埋めて確定"),
    (26, "配信設計⑥｜通知メッセージ", "意見", "緑", "利用シーンの設計", "通知メッセージの申請可否・要件を確認"),
    (27, "改善①｜LTVを決めているのは、解約率（★山）", "事実", "緑", "継続期間（3%→約33ヶ月／7%→約14ヶ月）・チャーン削減幅は定期通販CRM各社の公開値。初回離脱率・F2転換率は業界水準", "―"),
    (28, "改善②｜工数削減と改善モデル", "意見", "緑", "DYM提案の改善モデル。人件費の時給換算では語っていない", "―"),
    (29, "成果｜効果測定の設計", "意見", "緑", "数値は入れない（業界汎用のためSIMは作らない）。限界CPOの式は業界標準", "御社の実績をいただければSIMを作成"),
    (30, "費用プラン", "意見", "緑", "DYM標準プラン（6ヶ月〜・税抜）。継続費を明記", "―"),
    (31, "運用スケジュール（6ヶ月）", "意見", "緑", "DYM標準の運用スケジュール", "ID連携の可否確認を最優先"),
    (32, "サポート体制", "意見", "緑", "DYM標準のサポート体制＋法令チェック体制", "―"),
    (33, "締め①｜飛び道具｜解約ボタンをLINEに置く", "意見", "緑", "DYM提案。法令上のメリットは特商法に基づく", "法務確認を推奨"),
    (34, "締め②｜第2の提案軸｜通知メッセージ", "意見", "緑", "DYM提案の第2軸", "通知メッセージの申請可否を確認"),
    (35, "締め③｜LINEOA実績", "事実", "黄", "LINE国内MAU・開封率はLINEヤフー公式値（緑相当）。D2C・単品通販の事例は★未取得", "lycbiz.comから事例取得。無ければ「該当なし」と正直に報告"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（原本のまま）", "―"),
]
assert len(ROWS) == 36, len(ROWS)

wb = Workbook()

# ---------------- サマリ ----------------
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "単品通販業界_LINEOA施策提案（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-09-02　対象ファイル：単品通販業界_LINEOA施策提案.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("単品通販のLTVを決めているのは、新規獲得ではなく解約率。月次解約率が3%なら顧客は33ヶ月、"
            "7%なら14ヶ月しか続かない。そして法律は「解約させない」を禁じた。引き止めるのではなく、"
            "解約の手前で受け止める導線をつくる。それができれば、限界CPOが上がって、広告が楽になる。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（未取得・仮説明記のうえ公表可）", "黄", Y_FILL),
    ("赤（出典確定まで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = '=IF(D11=0,"OK：赤0件。対外提出可（黄は未取得データの明記つきで提出可）","NG：赤"&D11&"件。出典を確定または該当数値を削除するまで対外提出不可")'
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 上司確認が必要な3点"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① DYM社内の単品通販広告実績（初回購入／定期引上／F2のCV地点別CPC・CVR・CPO）→ S8が実測になる",
    "② 薬機法・特商法で審査／表示に引っかかった実例（社内1〜2例）→ S9が一般論から実務に変わる",
    "③ 競合8社・LY公式事例の実名掲載OK（S12／S35）→ NGなら「A社／B社」表記に変更",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B22"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. Googleトレンド（サプリ／青汁／定期便／解約・一語・5年＋1年。スケールが違うKWは別グラフ）→ S10が黄→緑",
    "2. 上司確認3点セット（上記★）→ S8・S9・S12・S35が黄→緑",
    "3. LINEヤフーの前後検索データ（サプリ／定期便／解約／口コミ。★「解約」が本命）→ S11が黄→緑",
    "4. 競合の友だち数実測（本流8社。page.line.meから取得日つき。解約導線がLINEで完結するかも記録）→ S12が黄→緑",
    "5. LINEヤフー公式のD2C・単品通販導入事例（lycbiz.com）→ S35が黄→緑",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B29"] = "凡例（このファイルの使い方）"
ws["B29"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
]):
    ws.cell(30 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

# ---------------- 判定一覧 ----------------
ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 30

for col, w in zip("ABCDEF", [5, 42, 11, 7, 52, 46]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
