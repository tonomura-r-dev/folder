# -*- coding: utf-8 -*-
"""月次管理ブックから「既存アタック用の案件一覧」を別ファイルで作る。

  python _build/build_attack_list.py <月次管理ブック.xlsx> [出力先.xlsx]

佐村さんの指摘（2026-08-25）
  ・既存アタックをするのにエンドクライアント名が無いと使えない
  ・担当が分からないと誰が動くか決められない
  ・全体シート（月次管理ブック）にタブを足すのは怖いのでやめる → 別ファイルにする

シート構成
  エンド別サマリ : 1行=1エンドクライアント。売上の大きい順＝アタックの優先順位
  案件一覧       : 1行=1案件。エンド名・担当つきの明細。ここから電話をかける
  商材別サマリ   : 商材カテゴリ × 契約形態 × 売上 × 粗利（当初の依頼どおりの集計）

元ブックは読むだけで一切変更しない。
"""
import os
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string

FONT = "メイリオ"
NAVY = "1F3864"
SUB = "8496B0"
LIGHT = "D9E2F3"

YEN = "#,##0;-#,##0;\"-\""
PCT = "0.0%"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

MEDIA = ["AD", "アフィ", "CS", "MEO", "制作", "PR", "風評", "タレントシェア",
         "LINE公式アカウント", "メディア", "ベトナム", "ASP", "マス広告(MA)・その他"]

# 担当の列名。媒体によって呼び方が違う（ASPだけ広告主担当／媒体担当）
ROLE_NAMES = ["アカウント", "コンサル", "運用①", "運用②", "運用③",
              "広告主担当①", "広告主担当②", "媒体担当①", "媒体担当②"]

KEIS = ["ストック", "ショット"]


def f(sz=10, bold=False, color="000000"):
    return Font(name=FONT, size=sz, bold=bold, color=color)


def norm(v):
    return "" if v is None else str(v).strip()


def num(v):
    return v if isinstance(v, (int, float)) else 0


def header_map(ws):
    """1行目の見出しから {見出し: 0始まりの列番号} を作る。"""
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
    out = {}
    for i, v in enumerate(row):
        name = norm(v).replace("\n", "")
        if name and name not in out:  # 同名列は左側（通常ぶん）を優先
            out[name] = i
    return out


def extract(src_path):
    """全媒体シートから1行=1案件のレコードを取り出す。"""
    wb = load_workbook(src_path, read_only=True, data_only=True)
    recs = []
    skipped = []
    for sheet in MEDIA:
        ws = wb[sheet]
        h = header_map(ws)
        need = ["案件名", "社名", "エンドクライアント名", "商材", "計上種別",
                "請求額（税抜）", "原価合計", "利益"]
        missing = [n for n in need if n not in h]
        if missing:
            skipped.append((sheet, missing))
            continue
        # ヨミ／請求の判別列（ADだけ持っている）
        yi = next((i for n, i in h.items() if "ヨミか請求" in n), None)
        roles = [(n, h[n]) for n in ROLE_NAMES if n in h]

        for r in ws.iter_rows(min_row=2, max_col=80, values_only=True):
            if r[0] is None:
                continue
            kei = norm(r[h["計上種別"]])
            uri = num(r[h["請求額（税抜）"]])
            if not kei and uri == 0:
                continue  # 空行
            tanto = [norm(r[i]) for _, i in roles]
            tanto = [t for t in tanto if t]
            yomi = ""
            if yi is not None:
                s = norm(r[yi])
                yomi = "ヨミ" if "ヨミ" in s else ("請求" if "請求" in s else "")
            recs.append({
                "媒体": sheet,
                "社名": norm(r[h["社名"]]),
                "エンドクライアント名": norm(r[h["エンドクライアント名"]]) or norm(r[h["社名"]]),
                "案件名": norm(r[h["案件名"]]),
                "種別": norm(r[h["種別"]]) if "種別" in h else norm(r[h.get("商流", 0)]),
                "商材カテゴリ": norm(r[h["商材"]]),
                "サービス詳細": norm(r[h["サービス詳細"]]) if "サービス詳細" in h else "",
                "契約形態": kei,
                "対象月": norm(r[h["対象月"]]) if "対象月" in h else "",
                "ヨミ/請求": yomi,
                "売上": uri,
                "原価": num(r[h["原価合計"]]),
                "粗利": num(r[h["利益"]]),
                "担当": " / ".join(dict.fromkeys(tanto)),  # 重複を除いて連結
                "主担当": tanto[0] if tanto else "",
            })
    wb.close()
    return recs, skipped


# ============================================================ 共通の書式
def write_table(ws, headers, rows, widths, money_cols=(), pct_cols=(),
                start=1, wrap_cols=()):
    for i, hname in enumerate(headers, start=1):
        c = ws.cell(row=start, column=i, value=hname)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[start].height = 30

    for j, row in enumerate(rows):
        rr = start + 1 + j
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=rr, column=i, value=v)
            c.font = f(9)
            c.border = BORDER
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(i in wrap_cols))
        for i in money_cols:
            ws.cell(row=rr, column=i).number_format = YEN
        for i in pct_cols:
            ws.cell(row=rr, column=i).number_format = PCT

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    ws.auto_filter.ref = f"A{start}:{get_column_letter(len(headers))}{start + len(rows)}"


# ============================================================ ① エンド別サマリ
def build_end_summary(wb, recs):
    ws = wb.create_sheet("エンド別サマリ")
    g = defaultdict(lambda: {"件数": 0, "売上": 0.0, "粗利": 0.0,
                             "媒体": [], "担当": [], "商材": []})
    for r in recs:
        k = r["エンドクライアント名"] or "（エンド名なし）"
        d = g[k]
        d["件数"] += 1
        d["売上"] += r["売上"]
        d["粗利"] += r["粗利"]
        for key, col in (("媒体", "媒体"), ("商材", "商材カテゴリ")):
            if r[col] and r[col] not in d[key]:
                d[key].append(r[col])
        for t in r["担当"].split(" / "):
            if t and t not in d["担当"]:
                d["担当"].append(t)

    order = sorted(g, key=lambda k: -g[k]["売上"])
    rows = []
    for i, k in enumerate(order, start=1):
        d = g[k]
        rows.append([i, k, d["件数"], d["売上"], d["粗利"],
                     (d["粗利"] / d["売上"] if d["売上"] else None),
                     " / ".join(d["媒体"]), " / ".join(d["商材"]), " / ".join(d["担当"])])

    write_table(
        ws,
        ["No.", "エンドクライアント名", "案件数", "売上", "粗利", "粗利率",
         "媒体", "商材カテゴリ", "担当"],
        rows,
        [6, 34, 8, 14, 13, 9, 22, 26, 30],
        money_cols=(4, 5), pct_cols=(6,), start=3, wrap_cols=(7, 8, 9))

    ws["A1"] = "エンド別サマリ｜既存アタックの優先順位（売上の大きい順）"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = (f"{len(order)}社／このリストの上から当たるのが基本。"
                "担当欄の人に事前に一声かけてから動くこと。")
    ws["A2"].font = f(9, color="44546A")
    return ws


# ============================================================ ② 案件一覧
def build_detail(wb, recs):
    ws = wb.create_sheet("案件一覧")
    # エンドの売上規模が大きい順 → 同じエンドの中は売上降順
    tot = defaultdict(float)
    for r in recs:
        tot[r["エンドクライアント名"]] += r["売上"]
    ordered = sorted(recs, key=lambda r: (-tot[r["エンドクライアント名"]],
                                          r["エンドクライアント名"], -r["売上"]))
    cols = ["媒体", "社名", "エンドクライアント名", "案件名", "種別", "商材カテゴリ",
            "サービス詳細", "契約形態", "対象月", "ヨミ/請求", "売上", "原価", "粗利"]
    rows = []
    for r in ordered:
        row = [r[c] for c in cols]
        row.append(r["粗利"] / r["売上"] if r["売上"] else None)
        row.append(r["担当"])
        rows.append(row)

    write_table(
        ws, cols + ["粗利率", "担当（アカウント／コンサル／運用）"], rows,
        [14, 24, 24, 34, 10, 16, 16, 10, 12, 10, 14, 13, 13, 9, 30],
        money_cols=(11, 12, 13), pct_cols=(14,), start=3, wrap_cols=(4, 15))

    ws["A1"] = "案件一覧｜1行=1案件（エンド名・担当つき）"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = (f"{len(rows)}件／金額は税抜／エンドの売上が大きい順に並べています。"
                "オートフィルタで媒体・契約形態・担当で絞れます。")
    ws["A2"].font = f(9, color="44546A")
    return ws


# ============================================================ ③ 商材別サマリ
def build_cat_summary(wb, recs):
    ws = wb.create_sheet("商材別サマリ")
    g = defaultdict(lambda: [0, 0.0, 0.0])
    for r in recs:
        k = (r["媒体"], r["商材カテゴリ"] or "（未設定）", r["契約形態"] or "（未設定）")
        g[k][0] += 1
        g[k][1] += r["売上"]
        g[k][2] += r["粗利"]
    order = sorted(g, key=lambda k: (MEDIA.index(k[0]) if k[0] in MEDIA else 99,
                                     -g[k][1]))
    rows = [[m, c, k, v[0], v[1], v[2], (v[2] / v[1] if v[1] else None)]
            for (m, c, k) in order for v in [g[(m, c, k)]]]

    n = len(recs)
    s = sum(r["売上"] for r in recs)
    p = sum(r["粗利"] for r in recs)
    rows.append(["合計", "", "", n, s, p, (p / s if s else None)])
    for kei in KEIS:
        sel = [r for r in recs if r["契約形態"] == kei]
        ks = sum(r["売上"] for r in sel)
        kp = sum(r["粗利"] for r in sel)
        rows.append([f"　うち{kei}", "", "", len(sel), ks, kp, (kp / ks if ks else None)])

    write_table(
        ws, ["媒体", "商材カテゴリ", "契約形態", "件数", "売上", "粗利", "粗利率"],
        rows, [20, 20, 10, 8, 16, 14, 9],
        money_cols=(5, 6), pct_cols=(7,), start=3)

    last = 3 + len(rows)
    for rr in range(last - 2, last + 1):
        for c in range(1, 8):
            ws.cell(row=rr, column=c).font = f(10, bold=True, color=NAVY)
            ws.cell(row=rr, column=c).fill = PatternFill("solid", fgColor=LIGHT)

    ws["A1"] = "商材別サマリ｜商材カテゴリ × 契約形態 × 売上 × 粗利"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = "当初のご依頼どおりの集計。合計は案件一覧と一致します。"
    ws["A2"].font = f(9, color="44546A")
    return ws


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "既存アタック用_案件一覧.xlsx"

    recs, skipped = extract(src)
    wb = Workbook()
    wb.remove(wb.active)
    build_end_summary(wb, recs)
    build_detail(wb, recs)
    build_cat_summary(wb, recs)
    wb.save(out)

    ends = {r["エンドクライアント名"] for r in recs}
    no_end = sum(1 for r in recs if not r["エンドクライアント名"])
    no_tanto = sum(1 for r in recs if not r["担当"])
    print(f"書き出しました: {out}")
    print(f"  案件数        : {len(recs):,}")
    print(f"  エンド社数    : {len(ends):,}")
    print(f"  売上          : {sum(r['売上'] for r in recs):,.0f}")
    print(f"  粗利          : {sum(r['粗利'] for r in recs):,.0f}")
    print(f"  エンド名なし  : {no_end}件")
    print(f"  担当なし      : {no_tanto}件")
    for sheet, miss in skipped:
        print(f"  ★スキップ {sheet}: {miss}")


if __name__ == "__main__":
    main()
