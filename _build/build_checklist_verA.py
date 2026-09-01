# -*- coding: utf-8 -*-
"""賃貸業界 Ver.A（入居者集客版）資料 公表判断チェックリスト（xlsx）

  python _build/build_checklist_verA.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説。データが入れば緑になる（公表可だが「未取得」明記が必要）
  赤 = 出典未記載の数値が載っている（出典を確定 or 削除するまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "賃貸業界資料_公表判断チェックリスト_VerA.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")
Y_FILL = PatternFill("solid", fgColor="FFF2CC")
R_FILL = PatternFill("solid", fgColor="FCE4E4")
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "検討背景", "事実", "緑", "LINEヤフー公式（国内MAU・開封率）", "―"),
    (4, "賃貸仲介のWebマーケ構造｜3つの壁", "意見", "緑", "DYM提案の構造整理。数値なし", "―"),
    (5, "部屋探しユーザーの心理", "事実", "緑", "SUUMOリサーチセンター 賃貸契約者動向調査（2022〜2024年度・首都圏）", "検索スニペット経由のため提案前に一次ソースで再確認"),
    (6, "できること／できないこと＋スコープ宣言", "意見", "緑", "DYM提案のスコープ宣言。数値なし", "―"),
    (7, "ニーズ調査｜シーズナリティ", "事実", "緑", "Googleトレンド実データ（2026年8月取得）", "提案時に取得日を更新"),
    (8, "ニーズ調査｜前後検索", "事実", "緑", "LINEヤフー媒体資料（取得済み・反映済み）", "―"),
    (9, "他社分析①｜競合のLINE運用ステータス", "事実", "黄", "2社選定・IDは事実。友だち数は★未取得", "page.line.meから取得日つきで実測"),
    (10, "他社分析②｜大手FCは店舗ごとにアカウントが乱立", "事実", "黄", "店舗別アカウントの存在は確認済み。友だち数は★未取得", "page.line.meから店舗ごとに実測"),
    (11, "全体設計｜カスタマージャーニー", "意見", "緑", "DYM標準フレームに賃貸仲介を当てはめ", "―"),
    (12, "全体設計｜施策全体像", "意見", "緑", "DYM提案の施策全体像", "―"),
    (13, "全体設計｜対策領域マップ", "意見", "緑", "新規反響・既存入居者の2領域", "―"),
    (14, "全体設計｜施策展開図（初期・月次）", "意見", "緑", "DYM標準の展開図を賃貸施策に置換", "―"),
    (15, "構築｜友だち追加動線", "意見", "緑", "DYM提案の動線設計", "―"),
    (16, "構築｜あいさつメッセージ", "意見", "緑", "配信文面は初稿", "担当者名・実際の運用フローに差し替え"),
    (17, "構築｜希望条件アンケート", "意見", "緑", "アンケート設計。数値なし", "―"),
    (18, "構築｜リッチメニュー", "意見", "緑", "DYM提案のリッチメニュー設計", "―"),
    (19, "構築｜反響即時対応", "意見", "緑", "自動化設計の説明。数値なし", "―"),
    (20, "配信設計｜14日ステップ設計", "意見", "緑", "配信設計。日数はS08前後検索データと連動", "―"),
    (21, "配信設計｜実文面①", "意見", "緑", "配信文面は初稿", "実際の物件情報に差し替え"),
    (22, "配信設計｜実文面②", "意見", "緑", "配信文面は初稿", "―"),
    (23, "配信設計｜物件レコメンド配信", "意見", "緑", "配信設計。S17のタグ設計が前提", "―"),
    (24, "配信設計｜年間企画投稿カレンダー", "推定", "緑", "S07のGoogleトレンド実データに基づく企画案", "―"),
    (25, "配信設計｜通知メッセージ", "意見", "緑", "利用シーンの設計", "電話番号・LINE ID取得の同意状況を確認"),
    (26, "来店改善・工数｜来店率改善", "事実", "緑", "オンライン内見実施率32.5%（SUUMOリサーチセンター）", "―"),
    (27, "来店改善・工数｜工数削減", "意見", "緑", "DYM提案の工数削減設計", "―"),
    (28, "来店改善・工数｜改善モデル", "意見", "緑", "DYM提案の改善モデル", "―"),
    (29, "成果と体制｜効果測定の設計", "意見", "緑", "数値は入れない（業界汎用のためSIMは作らない）", "御社の実績をいただければSIMを作成"),
    (30, "成果と体制｜費用プラン", "意見", "緑", "DYM標準プラン（6ヶ月〜・税抜）", "―"),
    (31, "成果と体制｜運用スケジュール", "意見", "緑", "DYM標準の運用スケジュール", "―"),
    (32, "成果と体制｜サポート体制", "意見", "緑", "DYM標準のサポート体制", "―"),
    (33, "締め｜飛び道具", "事実", "緑", "D案は前後検索データ（S08）で裏付け済み。A〜Cは意見（施策アイデア）", "―"),
    (34, "締め｜第2の提案軸｜オーナー向け", "意見", "緑", "Ver.Bへの橋渡し。数値なし", "―"),
    (35, "締め｜LINEOA実績", "事実", "緑", "ハートサポート／LIFULL HOME'S（LINEヤフー公式事例）", "検索スニペット経由のため掲載前に各事例ページの原典で再確認"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（原本のまま）", "―"),
]
assert len(ROWS) == 36, len(ROWS)

wb = Workbook()
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "賃貸業界_LINEOA施策提案_VerA_入居者集客（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-31　対象ファイル：賃貸業界_LINEOA施策提案_VerA_入居者集客.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("ポータル反響は、いま「使い捨て」になっている。Webマーケでできるのは、"
            "反響を資産に変えてCPOを下げること。接客の質ではなく、反響の再利用率で勝つ。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（未取得明記のうえ公表可）", "黄", Y_FILL),
    ("赤（出典確定まで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = '=IF(D11=0,"OK：赤0件。対外提出可（黄は未取得データの明記つきで提出可）","NG：赤"&D11&"件。出典を確定または該当数値を削除するまで対外提出不可")'
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 上司確認が必要な3点"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① 社内実績の共有可否（本資料はSIMを作らない構成のため必須ではないが、次のSIM作成時に必要）",
    "② SUUMOリサーチセンター・日管協の一次ソース再確認（検索スニペット経由の数値。S05）",
    "③ 競合2社・大手FCの実名掲載OK（S09／S10）→ NGなら「A社／B社」表記に変更",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B22"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. Googleトレンド（取得済み。スケール分けの再取得のみ。S07）",
    "2. 上司確認3点セット（上記★）→ S05・S09・S10が黄→緑",
    "3. LINEヤフーの前後検索データ（取得済み・反映済み。S08）",
    "4. 競合の友だち数実測（いい部屋ネット本部／エイブルAGENT／大手FC各店舗。管理併営かも記録）→ S09・S10が黄→緑",
    "5. 一次ソースの裏取り（SUUMOリサーチセンター／LY事例ページ）→ S05・S35",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B29"] = "凡例（このファイルの使い方）"
ws["B29"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
]):
    ws.cell(30 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 30

for col, w in zip("ABCDEF", [5, 42, 11, 7, 52, 46]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
