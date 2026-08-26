# -*- coding: utf-8 -*-
"""CRMスプシに数字を移すための「引き当て用マスタ」を作る。

  python _build/build_crm_master.py <月次管理ブック.xlsx> [出力先ベース名]

CRM側の列構成が分からなくても使えるように、社名キーとエンド名キーの
両方でマスタを出す。CRM側から VLOOKUP / XLOOKUP で引けばよい。

出力
  <ベース名>.xlsx        : エンド名キー / 社名キー の2シート
  <ベース名>_エンド名.tsv : CRMの新規タブに貼る用（エンド名キー）
  <ベース名>_社名.tsv     : 同上（社名キー）

契約形態は1社が両方持つことがあるので「ストック／ショット／ストック＋ショット」の
3値にし、金額もストックぶんとショットぶんに分ける。

元ブックは読むだけで一切変更しない。
"""
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_attack_list import MEDIA, extract  # 抽出処理は共通

FONT = "メイリオ"
NAVY = "1F3864"
YEN = "#,##0;-#,##0;\"-\""
PCT = "0.0%"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

HEADERS = ["キー", "契約形態", "ストック売上", "ショット売上", "売上計",
           "粗利計", "粗利率", "案件数", "媒体", "商材カテゴリ", "担当", "最新対象月"]


def f(sz=10, bold=False, color="000000"):
    return Font(name=FONT, size=sz, bold=bold, color=color)


def month_key(s):
    """「8月分」などから並べ替え用の数値を作る。判別できないものは0。"""
    digits = ""
    for ch in s:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    return int(digits) if digits else 0


def aggregate(recs, key_field):
    """key_field（社名 or エンドクライアント名）ごとに1行にまとめる。"""
    g = defaultdict(lambda: {"ストック": 0.0, "ショット": 0.0, "売上": 0.0,
                             "粗利": 0.0, "件数": 0, "媒体": [], "商材": [],
                             "担当": [], "月": [], "形態": set()})
    for r in recs:
        k = r[key_field]
        if not k:
            continue
        d = g[k]
        d["件数"] += 1
        d["売上"] += r["売上"]
        d["粗利"] += r["粗利"]
        if r["契約形態"] in ("ストック", "ショット"):
            d[r["契約形態"]] += r["売上"]
            # 金額が0でも契約形態は判定できるので、種別そのものを持っておく
            d["形態"].add(r["契約形態"])
        for src, dst in (("媒体", "媒体"), ("商材カテゴリ", "商材")):
            if r[src] and r[src] not in d[dst]:
                d[dst].append(r[src])
        for t in r["担当"].split(" / "):
            if t and t not in d["担当"]:
                d["担当"].append(t)
        if r["対象月"] and r["対象月"] not in d["月"]:
            d["月"].append(r["対象月"])

    rows = []
    for k in sorted(g, key=lambda x: -g[x]["売上"]):
        d = g[k]
        # 金額ではなく計上種別そのもので判定する（売上0の案件も拾うため）
        has_s = "ストック" in d["形態"]
        has_o = "ショット" in d["形態"]
        if has_s and has_o:
            keiyaku = "ストック＋ショット"
        elif has_s:
            keiyaku = "ストック"
        elif has_o:
            keiyaku = "ショット"
        else:
            keiyaku = "（未設定）"
        latest = max(d["月"], key=month_key) if d["月"] else ""
        rows.append([
            k, keiyaku, d["ストック"], d["ショット"], d["売上"], d["粗利"],
            (d["粗利"] / d["売上"] if d["売上"] else None),
            d["件数"], " / ".join(d["媒体"]), " / ".join(d["商材"]),
            " / ".join(d["担当"]), latest,
        ])
    return rows


def write_sheet(wb, title, rows, key_label, note):
    ws = wb.create_sheet(title)
    ws["A1"] = f"{title}｜CRMから VLOOKUP で引く用"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = note
    ws["A2"].font = f(9, color="44546A")

    headers = [key_label] + HEADERS[1:]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 30

    for j, row in enumerate(rows):
        rr = 4 + j
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=rr, column=i, value=v)
            c.font = f(9)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=(i in (9, 10, 11)))
        for i in (3, 4, 5, 6):
            ws.cell(row=rr, column=i).number_format = YEN
        ws.cell(row=rr, column=7).number_format = PCT

    for i, w in enumerate([34, 16, 14, 14, 14, 13, 9, 8, 22, 26, 30, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{3 + len(rows)}"
    return ws


def write_tsv(path, rows, key_label):
    with open(path, "w", encoding="utf-8") as fh:
        fh.write("\t".join([key_label] + HEADERS[1:]) + "\n")
        for row in rows:
            out = []
            for i, v in enumerate(row):
                if v is None:
                    out.append("")
                elif i == 6:  # 粗利率だけ小数、金額は整数で出す
                    out.append(f"{v:.4f}")
                elif isinstance(v, float):
                    out.append(f"{v:.0f}")
                else:
                    out.append(str(v))
            fh.write("\t".join(out) + "\n")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    base = sys.argv[2] if len(sys.argv) > 2 else "CRM引き当てマスタ"

    recs, skipped = extract(src)
    end_rows = aggregate(recs, "エンドクライアント名")
    co_rows = aggregate(recs, "社名")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "エンド名キー", end_rows, "エンドクライアント名",
                f"{len(end_rows)}社／実際にアタックする相手の名前で引く場合はこちら。")
    write_sheet(wb, "社名キー", co_rows, "社名",
                f"{len(co_rows)}社／請求先（代理店含む）の名前で引く場合はこちら。")
    wb.save(f"{base}.xlsx")

    write_tsv(f"{base}_エンド名.tsv", end_rows, "エンドクライアント名")
    write_tsv(f"{base}_社名.tsv", co_rows, "社名")

    print(f"書き出しました: {base}.xlsx / {base}_エンド名.tsv / {base}_社名.tsv")
    print(f"  エンド名キー : {len(end_rows):,}社")
    print(f"  社名キー     : {len(co_rows):,}社")
    for label, rows in (("エンド名", end_rows), ("社名", co_rows)):
        c = defaultdict(int)
        for r in rows:
            c[r[1]] += 1
        print(f"  {label}の契約形態内訳: {dict(c)}")
    for sheet, miss in skipped:
        print(f"  ★スキップ {sheet}: {miss}")


if __name__ == "__main__":
    main()
