# -*- coding: utf-8 -*-
"""賃貸業界資料 Ver.B（オーナー開拓版）公表判断チェックリスト（xlsx）

  python _build/build_checklist_chintai_verB.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説、一次ソース未確認、または監修が必要（条件付きで公表可）
  赤 = 出典未記載の数値、または未取得の差込枠が残っている（埋まるまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DECK = ROOT / "賃貸業界_LINEOA施策提案_VerB_オーナー開拓.pptx"
OUT = ROOT / "賃貸業界資料_公表判断チェックリスト_VerB.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑
Y_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # 赤
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")  # 記入セル
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (枚, タイトル, 主張の種類, 判定, 根拠・出典, 残タスク)
ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "LINE公式アカウントを検討する背景", "事実", "黄",
     "LINE国内MAU 1億人突破（2025年12月末）／開封率 約55%（2022年6月時点）＝LINEヤフー公式。"
     "★「オーナーもLINEを使う」とは書かず、消去法で記述している",
     "★GMO調査に「オーナーが使いたいツールはメールが首位」という反証材料あり。"
     "根拠が出せないなら消去法のまま提出する"),
    (4, "市場環境｜今すぐ検討中の人しか拾えていない", "事実", "黄",
     "家賃滞納率＝日管協短観2022年度（検索スニペット経由）。検討期間・受注件数・単価はDYM見解",
     "日管協短観を一次ソースで再確認（第29回は発表時期未定）。"
     "販促費・面談数・受注件数を貴社実績に差し替え"),
    (5, "オーナー理解｜不満はある。でも替えるのが面倒", "意見", "黄",
     "「心の声」6つはDYM見解（調査データではない）。裏付けは前後検索データのみ",
     "心の声は仮説である旨を口頭で補足。可能なら貴社の失注理由データで検証"),
    (6, "本提案のスコープ｜できること／できないこと", "意見", "緑",
     "提案スコープの宣言。数値は記載していない", "―"),
    (7, "ニーズ調査（モーメント）", "事実", "緑",
     "Googleトレンド実測（日本／過去5年・2026年8月取得）。"
     "★「土地活用」「アパート経営」が測定不能であることも正直に記載済み",
     "提案直前に再取得して取得日を更新する"),
    (8, "ニーズ調査（前後検索）", "事実", "緑",
     "LINEヤフー 前後検索データ（Journey・直近1年／2026年8月受領）。起点KW「土地活用」を明記済",
     "―"),
    (9, "他社分析①｜オーナー向けアカウント", "事実", "赤",
     "★オーナー向けの別アカウントが1社も特定できておらず、友だち数も未取得。"
     "破線の差込枠のまま（クラウドから page.line.me がブロック）",
     "page.line.meで実測（PC作業）。オーナー向け別アカウントの有無と友だち数・取得日を記録。"
     "埋まるまで本スライドは対外提出しない"),
    (10, "他社分析②｜オーナー向けLINEは4型", "意見", "黄",
     "4分類はDYM見解（仮説バッジ表示済）。「オーナー向けアカウントが見つからない」は2026年8月の調査結果",
     "P9の実測後に各社をプロットして確定。競合の実名掲載可否を社内確認"),
    (11, "全体設計｜カスタマージャーニー × CV3段", "意見", "緑",
     "CV3段は本提案での定義（DYM設計）", "貴社のCV定義とすり合わせ"),
    (12, "施策全体像｜触るのは2箇所だけ", "意見", "緑",
     "スコープの図解。数値は記載していない", "―"),
    (13, "全体設計｜効果を最大化する2軸", "意見", "緑",
     "動線設計（DYM見解）。数値は記載していない", "経路別の実績値は運用2ヶ月目から実測"),
    (14, "LINE公式アカウントにおける対策領域", "意見", "緑",
     "機能マップ（FMT）。オーナー向けの文脈に置換済", "―"),
    (15, "施策展開図（初期・月次）", "意見", "緑",
     "実施メニュー（FMT）。「※有料」項目はP31と対応", "―"),
    (16, "構築｜友だち追加動線", "意見", "黄",
     "5経路の動線設計と実文面。数値は記載していない。"
     "①既存オーナーへの一斉登録は、管理委託契約の範囲内のご連絡として設計",
     "★既存オーナーへの登録案内について、契約範囲と個人情報の取得同意を法務・管理部門に確認"),
    (17, "構築｜あいさつメッセージ", "意見", "緑",
     "文面はDYM設計。数値の主張はない", "担当者名・対応時間を貴社の運用に合わせて差し替え"),
    (18, "構築｜土地活用タイプ診断", "意見", "黄",
     "設問設計。Q1を建築可否にする根拠はLINEヤフー前後検索データ（P8）",
     "★建築可否の判定は法令・条例（用途地域／市街化調整区域／農地転用）の解釈を含むため、"
     "貴社および顧問士業の監修が必須"),
    (19, "構築｜リッチメニュー 3タブ×6枠", "意見", "緑",
     "18ボタンの設計と実装仕様（2500×1686px）。数値の主張なし", "―"),
    (20, "構築｜長期育成の設計思想", "意見", "黄",
     "設計思想の宣言（DYM見解）。数値は記載していない。「検討1〜3年」を前提に置いている",
     "★「検討1〜3年」の前提を、貴社の受注実績（初回接触から受注までの期間）で検証する"),
    (21, "配信設計｜年次モーメントカレンダー", "事実", "緑",
     "参照データはGoogleトレンド実測とLINEヤフー前後検索データ。"
     "「土地活用」「アパート経営」は測定不能のため根拠に使っていない",
     "提案直前にGoogleトレンドを再取得して取得日を更新"),
    (22, "配信設計｜企画投稿（月次）", "意見", "緑",
     "配信テーマの設計。本資料に数値は記載していない",
     "エリア相場は貴社の管理物件・成約データから毎月作成する（データ提供が前提）"),
    (23, "配信設計｜初動30日のステップ", "意見", "緑",
     "各日の根拠はLINEヤフー前後検索データ（P8）", "配信間隔は運用開始後のブロック率で調整"),
    (24, "配信設計｜実文面（Day0／Day7／Day30）", "意見", "黄",
     "配信文面の初稿。★診断結果の数値（相場差 ▲8,000円/月 など）はすべて記入例であることを脚注に明記",
     "実運用では診断結果と貴社の実績事例を差し込む。事例に使う物件の掲載可否を確認"),
    (25, "配信設計｜セミナー送客", "意見", "緑",
     "送客フローの設計。参加率・予約率の数値は記載していない", "セミナーのテーマ例をP21に合わせて確定"),
    (26, "配信設計｜LINE通知メッセージ", "意見", "黄",
     "利用シーンの設計。管理委託契約に基づくご連絡として設計している",
     "★電話番号の取得同意状況を確認。営業目的の配信と分けて設計し、配信停止導線を用意する"),
    (27, "歩留まり｜相談の歩留まり改善", "意見", "緑",
     "施策設計。歩留まりの想定値は置いていない", "現在の予約→実施率、訪問1件あたりの移動時間を実測"),
    (28, "歩留まり｜工数削減", "意見", "緑",
     "Before/Afterの業務整理。削減時間の想定値は置いていない",
     "訪問件数・移動時間・架電件数の実測をもらって「営業◯名の◯日分」まで定量化"),
    (29, "改善モデル", "意見", "緑",
     "機能設計（FMT）。★FMT由来の出典未記載数値（CVR 2%→5〜10%）は本版で削除済", "―"),
    (30, "成果｜効果測定の設計", "意見", "緑",
     "測る段と主KPI（面談CPA・CVR）と構造のみ。★業界汎用のため数値は一切入れていない。"
     "CPOは3年累積の構造としてのみ提示し、約束していない",
     "貴社実績が入り次第、この構造に数字を入れて別紙SIMを作成"),
    (31, "体制｜費用プラン", "事実", "黄",
     "運用プラン・ツール費はDYM標準（2026年8月時点）。LINEヤフー社への支払分と分けて記載",
     "最新の料金表と一致しているか社内確認（上司確認②）"),
    (32, "体制｜運用スケジュール", "意見", "緑",
     "標準的な6ヶ月の進行", "既存アカウントの有無・ID連携・診断監修の所要期間で前後を調整"),
    (33, "体制｜サポート体制", "意見", "緑",
     "DYM標準の体制と分担", "有人チャットの対応時間・担当、診断監修の担当を運用開始前に取り決め"),
    (34, "締め｜飛び道具（4案）", "意見", "黄",
     "A・DはLINEヤフー前後検索データで裏付け済み。B・CはDYM見解（効果の裏付けなし）",
     "★Aのサブリース記述は、貴社の商品内容と法令表示（賃貸住宅管理業法・景表法）に照らして監修が必須"),
    (35, "締め｜LINEOA実績", "事実", "黄",
     "LINEヤフー公式が実名公開している事例4件。数値はスニペット経由。"
     "★「オーナー向けの公式事例は存在しない」も明記している",
     "掲載前に各事例ページ原典で数値を再確認。"
     "「存在しない」は2026年8月時点の調査結果であり不存在の証明ではない旨を口頭で補足"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（FMTのまま）", "―"),
]

wb = Workbook()

# ---------------- サマリ ----------------
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "賃貸業界_LINEOA施策提案 Ver.B｜オーナー開拓版（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-28　対象ファイル：賃貸業界_LINEOA施策提案_VerB_オーナー開拓.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("オーナーは営業されたくないが、情報は欲しい。しかも検討は1〜3年。"
            "訪問・電話・DMは「その瞬間に検討中の人」しか拾えない。"
            "やる意思はあるが今すぐではない層を育て続ければ、面談獲得の単価（CPA）が下がる。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（仮説明記・裏取り・監修のうえ公表可）", "黄", Y_FILL),
    ("赤（差込が埋まるまで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = ('=IF(D11=0,"OK：赤0件。対外提出可",'
             '"NG：赤"&D11&"件。差込枠が埋まるまで対外提出不可（該当スライドを抜けば提出可）")')
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ この版だけの注意点（Ver.Aには無いもの）"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・P18の診断（建築可否の判定）と、P34-Aのサブリース記述は、法令・条例・表示規制に関わります。"
    "社内の有資格者または顧問士業の監修を必ず受けてから提出してください。",
    "・P03は「オーナーもLINEを使う」と言い切っていません。"
    "GMO調査に「オーナーが使いたいツールはメールが首位」という反証材料があるためです。",
    "・FMT由来の出典未記載4数値（業態平均で最大3%／CVR 2%→5〜10%／商談化70%改善・再CV率3〜8%／"
    "友だち追加率0.8〜1.0%）は、本版では該当スライドの差し替えと文言修正で除去しています。",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 30

ws["B22"] = "★ 上司・社内で確認が必要な3点"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① 競合の実名掲載の可否（P9・P10）→ NGなら「A社／B社」表記に変更する。"
    "※LINEヤフーが実名公開している導入事例（P35）は出典明記で掲載可",
    "② 費用プラン（P31）の金額が最新のDYM料金表と一致しているか。ツール費・通知メッセージの通数単価も含む",
    "③ ★診断ロジック（P18）とサブリース記述（P34-A）の監修体制。誰が確認して出すかを決める",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 28

ws["B27"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B27"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. 競合の実測（★オーナー向けの別アカウントの有無／友だち数／取得日／診断コンテンツの有無）"
    "→ P9が赤→緑、P10が黄→緑",
    "2. 一次ソースの裏取り（日管協短観／LINEヤフー導入事例ページ）→ P4・P35が黄→緑",
    "3. ★オーナーのLINE利用実態の根拠（GMO調査の反証材料への対応）→ P3が黄→緑。"
    "根拠が出せない場合は消去法のまま提出する",
    "4. 貴社の実績値（販促費・資料請求数・訪問件数・面談数・受注件数・初回接触から受注までの期間）"
    "→ P4・P20・P28が実測になり、P30の構造に数字を入れて別紙SIMが作れる",
    "5. エリアの成約家賃データ（P22の月次配信の元データ）と、診断ロジックの監修（P18）",
]):
    r = 28 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 26

ws["B34"] = "凡例（このファイルの使い方）"
ws["B34"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
    "・本資料は業界汎用のため、成果シミュレーション（SIM）は入れていません（P30は構造のみ）。"
    "個社提案に落とすときに、上記4の実績値をもらってSIMを作ります。",
    "・主KPIはCPOではなく「相談・セミナー予約のCPA／友だち→予約CVR」です（Ver.Aとの最大の違い）。",
]):
    ws.cell(35 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

# ---------------- 判定一覧 ----------------
ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 42

for col, w in zip("ABCDEF", [5, 42, 11, 7, 58, 48]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
