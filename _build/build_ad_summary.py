# -*- coding: utf-8 -*-
"""月次管理ブックの「AD」シートから、商材カテゴリ×契約形態×売上×粗利×担当の
サマリを1枚にまとめた別ファイルを作る。元ブックは読むだけで一切変更しない。

  python _build/build_ad_summary.py <元ブック.xlsx> [出力先.xlsx]

出力
  サマリ : 見るのはこの1枚（検算／商材×契約形態／担当別／ヨミ・請求別）
  明細   : 上の集計の根拠。1行=1案件。ピボットの元データにも使える
"""
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC_SHEET = "AD"
FONT = "メイリオ"
NAVY = "1F3864"
SUB = "8496B0"
LIGHT = "D9E2F3"
WARN = "FFF2CC"
BLANK = "（未設定）"

YEN = "#,##0;-#,##0;\"-\""
PCT = "0.0%"
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 明細シートの列（元シートの列名をそのまま使う）
DETAIL_COLS = [
    ("No.", 1), ("ヨミ/請求", None), ("案件名", 3), ("種別", 4), ("社名", 5),
    ("エンドクライアント名", 6), ("商材", 16), ("サービス詳細", 17),
    ("計上種別", 11), ("対象月", 12),
    ("請求額（税抜）", 8), ("原価合計", 9), ("利益", 10), ("粗利率", None),
    ("アカウント", 18), ("コンサル", 19), ("運用①", 20), ("運用②", 21), ("運用③", 22),
    ("アカウント売上", 28), ("コンサル売上", 29), ("運用①売上", 30),
    ("運用②売上", 31), ("運用③売上", 32),
    ("アカウント利益", 33), ("コンサル利益", 34), ("運用①利益", 35),
    ("運用②利益", 36), ("運用③利益", 37),
]
ROLES = ["アカウント", "コンサル", "運用①", "運用②", "運用③"]
CAT_ORDER = ["リスティング", "SNS", "DSP", "その他"]
KEI_ORDER = ["ストック", "ショット"]


def f(sz=10, bold=False, color="000000", italic=False):
    return Font(name=FONT, size=sz, bold=bold, color=color, italic=italic)


def norm(v):
    return BLANK if v in (None, "") else str(v).strip()


def yomi_of(v):
    s = "" if v is None else str(v)
    if "ヨミ" in s:
        return "ヨミ"
    if "請求" in s:
        return "請求"
    return BLANK


def num(v):
    return v if isinstance(v, (int, float)) else None


# ============================================================ 抽出
def extract(src_path):
    wb = load_workbook(src_path, read_only=True, data_only=True)
    ws = wb[SRC_SHEET]
    rows = [r for r in ws.iter_rows(min_row=2, max_col=64, values_only=True)
            if r[0] is not None]
    out = []
    for r in rows:
        rec = {}
        for name, idx in DETAIL_COLS:
            if idx is None:
                continue
            rec[name] = r[idx - 1]
        rec["ヨミ/請求"] = yomi_of(r[1])
        for k in ("商材", "計上種別", "種別"):
            rec[k] = norm(rec[k])
        out.append(rec)
    wb.close()
    return out


def audit(recs):
    """元シートとの突合に使う実数と、データ品質の注意点を返す。"""
    s = sum(num(r["請求額（税抜）"]) or 0 for r in recs)
    c = sum(num(r["原価合計"]) or 0 for r in recs)
    p = sum(num(r["利益"]) or 0 for r in recs)
    # 按分売上の合計が本体売上と合っているか（担当別集計が信用できるかの判定）
    ap = sum((num(r[f"{role}売上"]) or 0) for r in recs for role in ROLES)
    gap = [r for r in recs
           if abs((num(r["請求額（税抜）"]) or 0)
                  - sum(num(r[f"{role}売上"]) or 0 for role in ROLES)) > 1]
    return {"件数": len(recs), "売上": s, "原価": c, "利益": p,
            "按分売上計": ap, "按分ズレ件数": len(gap)}


# ============================================================ 明細シート
def build_detail(wb, recs):
    ws = wb.create_sheet("明細")
    names = [n for n, _ in DETAIL_COLS]
    for i, n in enumerate(names, start=1):
        c = ws.cell(row=1, column=i, value=n)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[1].height = 32

    ci = {n: i + 1 for i, n in enumerate(names)}
    for j, r in enumerate(recs):
        row = j + 2
        for n in names:
            if n == "粗利率":
                ws.cell(row=row, column=ci[n],
                        value=f"=IF(N({get_column_letter(ci['請求額（税抜）'])}{row})=0,"
                              f"\"\",{get_column_letter(ci['利益'])}{row}"
                              f"/{get_column_letter(ci['請求額（税抜）'])}{row})")
            else:
                ws.cell(row=row, column=ci[n], value=r[n])
        for n in ("請求額（税抜）", "原価合計", "利益") + tuple(
                f"{role}{k}" for role in ROLES for k in ("売上", "利益")):
            ws.cell(row=row, column=ci[n]).number_format = YEN
        ws.cell(row=row, column=ci["粗利率"]).number_format = PCT
        for i in range(1, len(names) + 1):
            cell = ws.cell(row=row, column=i)
            cell.font = f(9)
            cell.border = BORDER

    widths = {"No.": 6, "ヨミ/請求": 10, "案件名": 34, "種別": 9, "社名": 26,
              "エンドクライアント名": 26, "商材": 13, "サービス詳細": 16,
              "計上種別": 10, "対象月": 12, "粗利率": 9}
    for n in names:
        ws.column_dimensions[get_column_letter(ci[n])].width = widths.get(n, 14)
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(names))}{len(recs) + 1}"
    return ci, len(recs) + 1


# ============================================================ サマリシート
def build_summary(wb, recs, ci, last, src_name, a):
    ws = wb.create_sheet("サマリ", 0)
    L = {n: get_column_letter(i) for n, i in ci.items()}
    D = f"明細!"

    def rng(col):
        return f"{D}${L[col]}$2:${L[col]}${last}"

    def head(row, text, note=""):
        ws.cell(row=row, column=1, value=text).font = f(12, bold=True, color=NAVY)
        if note:
            ws.cell(row=row, column=2, value=note).font = f(9, color="808080")

    def th(row, col, text, fill=NAVY):
        c = ws.cell(row=row, column=col, value=text)
        c.font = f(9, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        return c

    ws["A1"] = "AD案件サマリ｜商材カテゴリ × 契約形態 × 売上 × 粗利 × 担当"
    ws["A1"].font = f(14, bold=True, color=NAVY)
    ws["A2"] = (f"出典：{src_name} の「{SRC_SHEET}」シート（読み取りのみ・元ブックは未変更）／"
                f"{a['件数']}件／金額はすべて税抜")
    ws["A2"].font = f(9, color="44546A")
    ws.row_dimensions[1].height = 22

    # ---------- 検算
    head(4, "① 検算", "元のADシートと合っているかの確認。ここが一致していれば以下の集計も信用できます。")
    for i, t in enumerate(["項目", "本シート（自動計算）", "元ADシート（取込時の実数）", "差分"], start=1):
        th(5, i, t)
    checks = [
        ("件数", f"=COUNTA({rng('No.')})", a["件数"], "0"),
        ("売上（請求額・税抜）", f"=SUM({rng('請求額（税抜）')})", a["売上"], YEN),
        ("原価合計", f"=SUM({rng('原価合計')})", a["原価"], YEN),
        ("粗利（利益）", f"=SUM({rng('利益')})", a["利益"], YEN),
        ("担当按分売上の合計", f"=SUM({D}${L['アカウント売上']}$2:${L['運用③売上']}${last})",
         a["按分売上計"], YEN),
    ]
    for j, (label, formula, actual, fmt) in enumerate(checks):
        r = 6 + j
        ws.cell(row=r, column=1, value=label).font = f(10)
        ws.cell(row=r, column=2, value=formula).font = f(10, bold=True, color=NAVY)
        ws.cell(row=r, column=3, value=actual).font = f(10)
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").font = f(10, bold=True)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = BORDER
            if c > 1:
                ws.cell(row=r, column=c).number_format = fmt
    r = 11
    ws.cell(row=r, column=1,
            value=f"※担当按分にズレのある案件：{a['按分ズレ件数']}件"
                  + ("（按分割合の合計が100%でない行。担当別集計はこのぶんズレます）"
                     if a["按分ズレ件数"] else "（なし＝担当別集計と全体合計が一致）")
            ).font = f(9, color="C00000" if a["按分ズレ件数"] else "808080")

    # ---------- 商材 × 契約形態
    head(13, "② 商材カテゴリ × 契約形態", "契約形態＝元シートの「計上種別」（ストック／ショット）")
    cats = [c for c in CAT_ORDER if any(r["商材"] == c for r in recs)]
    cats += sorted({r["商材"] for r in recs} - set(cats))
    keis = [k for k in KEI_ORDER if any(r["計上種別"] == k for r in recs)]
    keis += sorted({r["計上種別"] for r in recs} - set(keis))

    ws.cell(row=14, column=1, value="商材カテゴリ")
    th(14, 1, "商材カテゴリ")
    ws.merge_cells(start_row=14, start_column=1, end_row=15, end_column=1)
    groups = keis + ["合計"]
    for g, name in enumerate(groups):
        base = 2 + g * 4
        th(14, base, name, NAVY if name != "合計" else SUB)
        ws.merge_cells(start_row=14, start_column=base, end_row=14, end_column=base + 3)
        for k, t in enumerate(["件数", "売上", "粗利", "粗利率"]):
            th(15, base + k, t, SUB)

    for i, cat in enumerate(cats + ["合計"]):
        r = 16 + i
        is_total = cat == "合計"
        ws.cell(row=r, column=1, value=cat).font = f(10, bold=is_total,
                                                     color=NAVY if is_total else "000000")
        for g, name in enumerate(groups):
            base = 2 + g * 4
            if is_total and name == "合計":
                cnt = f"=COUNTA({rng('No.')})"
                sal = f"=SUM({rng('請求額（税抜）')})"
                pro = f"=SUM({rng('利益')})"
            elif is_total:
                cnt = f"=COUNTIF({rng('計上種別')},\"{name}\")"
                sal = f"=SUMIF({rng('計上種別')},\"{name}\",{rng('請求額（税抜）')})"
                pro = f"=SUMIF({rng('計上種別')},\"{name}\",{rng('利益')})"
            elif name == "合計":
                cnt = f"=COUNTIF({rng('商材')},$A{r})"
                sal = f"=SUMIF({rng('商材')},$A{r},{rng('請求額（税抜）')})"
                pro = f"=SUMIF({rng('商材')},$A{r},{rng('利益')})"
            else:
                cnt = (f"=COUNTIFS({rng('商材')},$A{r},{rng('計上種別')},\"{name}\")")
                sal = (f"=SUMIFS({rng('請求額（税抜）')},{rng('商材')},$A{r},"
                       f"{rng('計上種別')},\"{name}\")")
                pro = (f"=SUMIFS({rng('利益')},{rng('商材')},$A{r},"
                       f"{rng('計上種別')},\"{name}\")")
            ws.cell(row=r, column=base, value=cnt).number_format = "0"
            ws.cell(row=r, column=base + 1, value=sal).number_format = YEN
            ws.cell(row=r, column=base + 2, value=pro).number_format = YEN
            sl, pl = get_column_letter(base + 1), get_column_letter(base + 2)
            ws.cell(row=r, column=base + 3,
                    value=f"=IF(N({sl}{r})=0,\"\",{pl}{r}/{sl}{r})").number_format = PCT
        for c in range(1, 2 + len(groups) * 4):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if not cell.font.bold:
                cell.font = f(10, bold=is_total)
            if is_total:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    # ---------- ヨミ / 請求
    top = 16 + len(cats) + 3
    head(top, "③ ヨミ／請求の内訳", "混ぜて見ないための内訳。②はヨミ＋請求の合計です。")
    for i, t in enumerate(["区分", "件数", "売上", "粗利", "粗利率"], start=1):
        th(top + 1, i, t)
    yomis = ["請求", "ヨミ"] + sorted({r["ヨミ/請求"] for r in recs} - {"請求", "ヨミ"})
    for i, y in enumerate(yomis + ["合計"]):
        r = top + 2 + i
        is_total = y == "合計"
        ws.cell(row=r, column=1, value=y)
        if is_total:
            ws.cell(row=r, column=2, value=f"=COUNTA({rng('No.')})")
            ws.cell(row=r, column=3, value=f"=SUM({rng('請求額（税抜）')})")
            ws.cell(row=r, column=4, value=f"=SUM({rng('利益')})")
        else:
            ws.cell(row=r, column=2, value=f"=COUNTIF({rng('ヨミ/請求')},$A{r})")
            ws.cell(row=r, column=3,
                    value=f"=SUMIF({rng('ヨミ/請求')},$A{r},{rng('請求額（税抜）')})")
            ws.cell(row=r, column=4, value=f"=SUMIF({rng('ヨミ/請求')},$A{r},{rng('利益')})")
        ws.cell(row=r, column=5, value=f"=IF(N(C{r})=0,\"\",D{r}/C{r})")
        ws.cell(row=r, column=2).number_format = "0"
        for c in (3, 4):
            ws.cell(row=r, column=c).number_format = YEN
        ws.cell(row=r, column=5).number_format = PCT
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = f(10, bold=is_total)
            if is_total:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    # ---------- 担当別
    top2 = top + 2 + len(yomis) + 3
    head(top2, "④ 担当別（按分後）",
         "1案件に複数担当がつくため、元シートの按分売上／按分利益で集計。二重計上なし。")
    for i, t in enumerate(["担当", "案件数（延べ）", "売上（按分）", "粗利（按分）", "粗利率"], start=1):
        th(top2 + 1, i, t)

    tan_sales = defaultdict(float)
    tan_cnt = Counter()
    for rec in recs:
        for role in ROLES:
            n = rec[role]
            if n in (None, ""):
                continue
            tan_sales[str(n).strip()] += num(rec[f"{role}売上"]) or 0
            tan_cnt[str(n).strip()] += 1
    order = sorted(tan_sales, key=lambda k: -tan_sales[k])

    for i, name in enumerate(order + ["合計"]):
        r = top2 + 2 + i
        is_total = name == "合計"
        ws.cell(row=r, column=1, value=name)
        if is_total:
            ws.cell(row=r, column=2,
                    value=f"=SUM(B{top2 + 2}:B{r - 1})")
            ws.cell(row=r, column=3,
                    value=f"=SUM({D}${L['アカウント売上']}$2:${L['運用③売上']}${last})")
            ws.cell(row=r, column=4,
                    value=f"=SUM({D}${L['アカウント利益']}$2:${L['運用③利益']}${last})")
        else:
            ws.cell(row=r, column=2, value="+".join(
                f"COUNTIF({rng(role)},$A{r})" for role in ROLES).join(("=", "")))
            ws.cell(row=r, column=3, value="+".join(
                f"SUMIF({rng(role)},$A{r},{rng(role + '売上')})" for role in ROLES
            ).join(("=", "")))
            ws.cell(row=r, column=4, value="+".join(
                f"SUMIF({rng(role)},$A{r},{rng(role + '利益')})" for role in ROLES
            ).join(("=", "")))
        ws.cell(row=r, column=5, value=f"=IF(N(C{r})=0,\"\",D{r}/C{r})")
        ws.cell(row=r, column=2).number_format = "0"
        for c in (3, 4):
            ws.cell(row=r, column=c).number_format = YEN
        ws.cell(row=r, column=5).number_format = PCT
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = f(10, bold=is_total)
            if is_total:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    # ---------- 注記
    n0 = top2 + 2 + len(order) + 3
    ws.cell(row=n0, column=1, value="読み方・注意点").font = f(12, bold=True, color=NAVY)
    notes = [
        "・商材カテゴリ＝元シートの「商材」列。契約形態＝「計上種別」列（ストック／ショット）。",
        "・売上＝「請求額（税抜）」、粗利＝「利益」。粗利率＝粗利÷売上。",
        "・②は対象月をまたいだ全行の合計です。月で絞るときは明細シートの「対象月」でフィルタ。",
        "・③のとおりヨミ（見込）と請求（確定）が同じ表に入っています。確定値だけ見たいときは「請求」の行を。",
        "・④は1案件を複数担当で分けた按分後の数字。①の「担当按分売上の合計」が売上と一致していれば正しく割れています。",
        "・元ブックは読み取りのみ。このファイルは別ファイルなので、元シートには一切影響しません。",
    ]
    for i, t in enumerate(notes):
        c = ws.cell(row=n0 + 1 + i, column=1, value=t)
        c.font = f(10)
        ws.merge_cells(start_row=n0 + 1 + i, start_column=1, end_row=n0 + 1 + i, end_column=8)

    ws.column_dimensions["A"].width = 24
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.sheet_view.showGridLines = False
    return ws


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "2026年8月_AD案件サマリ.xlsx"
    import os
    recs = extract(src)
    a = audit(recs)
    wb = Workbook()
    wb.remove(wb.active)
    ci, last = build_detail(wb, recs)
    build_summary(wb, recs, ci, last, os.path.basename(src), a)
    wb.save(out)
    print(f"書き出しました: {out}")
    for k, v in a.items():
        print(f"  {k}: {v:,}" if isinstance(v, (int, float)) else f"  {k}: {v}")


if __name__ == "__main__":
    main()
