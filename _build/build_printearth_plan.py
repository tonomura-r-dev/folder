#!/usr/bin/env python3
"""プリントアース LINE OA 9月配信企画案 → 配信カレンダー用 XLSX を生成する。

既存の配信カレンダーの列構成に合わせつつ、殿村さん指定の必須項目
（企画案 / 狙い概要 / ターゲット / バナー・構成案 / テキスト）を列として持たせる。

    python3 _build/build_printearth_plan.py

出力: プリントアース_LINEOA_配信企画案_2026-09.xlsx（リポジトリ直下）
文言の修正はこのファイルの PLANS を触る。企画の背景は
_drafts/プリントアース_LINEOA_配信企画案_2026-09.md を参照。
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = "https://www.inkart.jp"
UTM_SOURCE = "LINEOA"
UTM_MEDIUM = "crm"

# ブランドカラー（過去CR素材から抽出）
GREEN = "0E8A4F"
RED = "E8564B"
GRAY = "F2F2F2"

COLUMNS = [
    ("No", 5),
    ("配信日", 11),
    ("入稿日", 11),
    ("ステータス", 11),
    ("配信種別", 10),
    ("企画案", 26),
    ("狙い概要", 52),
    ("目的", 13),
    ("ターゲット", 24),
    ("配信対象", 22),
    ("画像サイズ", 12),
    ("バナー・構成案", 52),
    ("テキスト", 46),
    ("デザイン備考", 26),
    ("遷移先URL", 40),
    ("utm_source", 11),
    ("utm_medium", 11),
    ("utm_campaign", 32),
    ("最終URL", 56),
]

PLANS = [
    {
        "no": 21,
        "date": "2026/09/01",
        "draft": "2026/08/31",
        "kind": "通常配信",
        "title": "下期スタート・名刺",
        "aim": (
            "10/1の期初という「実在する締切」を先取りする。価格ではなくタイミングで"
            "CVさせるため、値引きを主役にしない。単価は低いがリピート率が高く、"
            "会員登録の入口として優秀。"
        ),
        "purpose": "CV獲得",
        "target": "BtoB・店舗・法人（検索広告流入層）",
        "audience": "全員",
        "banner": (
            "白フレーム＋緑見出し。黄吹き出しに「10/1に間に合う入稿締切◯/◯」。"
            "中央に名刺の実物写真（新春キャンペーンの商品写真ブロックを流用）。"
            "下部に赤CTA帯「名刺を注文する」。"
        ),
        "text": (
            "名刺、下期の分は足りていますか。\n\n"
            "10月1日に間に合わせるなら、\n"
            "入稿は9月◯日までが安心です。\n\n"
            "100枚◯◯◯円から、最短当日出荷。\n"
            "ロゴや肩書きの変更だけなら、\n"
            "前回のデータを直して出せます。\n\n"
            "ご注文はこちら（CTA）"
        ),
        "note": "※名刺カテゴリの実URLと価格を要確認",
        "url": f"{BASE}/lineup/",
        "campaign": "20260901_message_MeishiH2",
    },
    {
        "no": 22,
        "date": "2026/09/04",
        "draft": "2026/09/03",
        "kind": "セグメント配信",
        "title": "会員登録リマインド（1,000Pt）",
        "aim": (
            "現行#14の作り直し。「最大10万円分クーポン」を廃止し、広告で約束した"
            "1,000Ptに一本化する。全員ではなく会員登録リンク未クリック層に絞り、"
            "既登録者への無駄配信とブロックを止める。"
        ),
        "purpose": "CPF流入ユーザーの新規会員登録",
        "target": "CPF・広告流入で未だ会員登録していない友だち",
        "audience": "会員登録URL未クリック（オーディエンス除外）",
        "banner": (
            "現行#14のバナーを流用しつつ情報を削る。「最大10万円分」→「1,000円分の"
            "ポイント」に差し替え、「LINEスタンプカード併用」の一文は削除。"
            "黄吹き出し「30秒で完了」、赤CTA帯「ポイントを受け取る」。"
        ),
        "text": (
            "1,000円分のポイント、\n"
            "まだ受け取られていません。\n\n"
            "お手続きは会員登録だけ。\n"
            "30秒で終わります。\n\n"
            "次のご注文から使えるので、\n"
            "先に受け取っておいてください。\n\n"
            "受け取りはこちら（CTA）"
        ),
        "note": "※ポイントが初回注文から使える場合は「そのまま初回のお会計から引けます」に差し替え",
        "url": f"{BASE}/entry/regist/",
        "campaign": "20260904_message_SignUp1000pt",
    },
    {
        "no": 23,
        "date": "2026/09/08",
        "draft": "2026/09/07",
        "kind": "通常配信",
        "title": "秋の販促チラシ・最短当日出荷",
        "aim": (
            "検索広告の最大流入KW「チラシ印刷」「格安印刷通販」と同じ商材をLINEでも当てる。"
            "価格ではなくスピード（最短当日出荷）を主役にして、価格比較の消耗戦から降りる。"
        ),
        "purpose": "CV獲得",
        "target": "BtoB・店舗・販促担当（検索広告流入層）",
        "audience": "全員",
        "banner": (
            "年度末応援キャンペーンのレイアウト踏襲。背景を秋モチーフ（紅葉・オレンジ〜黄）に。"
            "左上に黄丸で「最短当日出荷」。中央白抜き帯に「チラシ・フライヤー」「ポスター」"
            "「ポスティング用」。下部にCTA帯。"
        ),
        "text": (
            "秋のセール、まだチラシ間に合います。\n\n"
            "A4チラシ1,000枚、\n"
            "データがあれば最短で当日出荷。\n\n"
            "サイズと部数を選ぶだけで、\n"
            "送料込みの総額が30秒で出ます。\n\n"
            "まず金額を見る（CTA）"
        ),
        "note": "※季節性のあるデザインでお願いします",
        "url": f"{BASE}/lineup/chirasi_flyer/",
        "campaign": "20260908_message_AutumnFlyer",
    },
    {
        "no": 24,
        "date": "2026/09/11",
        "draft": "2026/09/10",
        "kind": "通常配信",
        "title": "紙の見本帳、無料",
        "aim": (
            "本企画で最もCPA効率が良い一本。請求に会員登録が必要なので"
            "「請求→会員登録→ポイント発生→後日注文」が1本につながる。"
            "その場で買わない層をブロックさせずにCV（問い合わせ完了）で拾える。"
        ),
        "purpose": "CV獲得",
        "target": "全員（とくに未購入・検討中の層）",
        "audience": "全員",
        "banner": (
            "注文ガイドカードのフォーマット踏襲（白フレーム＋角の三角＋黄吹き出し＋緑特大見出し）。"
            "黄吹き出し「送料も無料」、見出し「紙の見本帳」。中央に見本帳の実物写真。"
            "赤CTA帯「無料で請求する」。"
        ),
        "text": (
            "紙は、画面では選べません。\n\n"
            "プリントアースの見本帳を、\n"
            "無料でお送りします。送料もかかりません。\n\n"
            "実際に印刷された用紙を手にとって、\n"
            "厚みも手ざわりも確かめてから決められます。\n\n"
            "無料で請求する（CTA）"
        ),
        "note": "※サイトイメージに合わせていただけると",
        "url": f"{BASE}/guide/request-info/",
        "campaign": "20260911_message_PaperSample",
    },
    {
        "no": 25,
        "date": "2026/09/16",
        "draft": "2026/09/15",
        "kind": "通常配信",
        "title": "2027年カレンダー 早割",
        "aim": (
            "9月の最大単価商材。ノベルティ需要は9〜10月に意思決定が動き、11月には"
            "他社で確定してしまうので先に取る。卓上は30個からで小規模店舗も対象になり"
            "母数を広く取れる。"
        ),
        "purpose": "CV獲得",
        "target": "BtoB全般（法人・士業・店舗・不動産・美容室などノベルティ配布業種）",
        "audience": "全員",
        "banner": (
            "キャンペーン型レイアウト。背景は冬の予感（濃紺〜白、雪の結晶を控えめに）。"
            "左上に黄丸で「早割」。中央に卓上（再生紙ケース／プラケース）と壁掛けの3点を"
            "横並び写真。右下に「〜◯/◯」の期限。"
        ),
        "text": (
            "2027年のカレンダー、\n"
            "もう動き出す時期です。\n\n"
            "卓上は30個から、\n"
            "壁掛けはA3・B4の大きいサイズまで。\n"
            "社名を入れて配れば1年間残ります。\n\n"
            "いま決めると早割対象です。\n"
            "種類と価格を見る（CTA）"
        ),
        "note": "※早割の割引率と期限を要確認",
        "url": f"{BASE}/lineup/calendar/",
        "campaign": "20260916_message_Calendar2027",
    },
    {
        "no": 26,
        "date": "2026/09/18",
        "draft": "2026/09/17",
        "kind": "通常配信",
        "title": "学園祭・文化祭",
        "aim": (
            "10〜11月の学園祭は締切が絶対的で、価格より納期で決まる。BtoBとBtoCの"
            "中間層（学生団体・サークル）が、うちわ・チラシ・パンフ・グッズをまとめて"
            "発注するため、1回のCVで点数が多い。"
        ),
        "purpose": "CV獲得",
        "target": "学生団体・サークル・PTA・地域イベント主催者",
        "audience": "全員",
        "banner": (
            "紙製うちわバナーのポップな路線を踏襲。実写＋明るいピンク／黄。"
            "「うちわ」「チラシ」「パンフレット」「ステッカー」の4点をアイコンで並列。"
            "下部CTA帯「まとめて見る」。"
        ),
        "text": (
            "学園祭の準備、印刷はお済みですか。\n\n"
            "うちわ、チラシ、パンフレット、ステッカー。\n"
            "プリントアースならまとめて刷れます。\n\n"
            "データさえあれば最短当日出荷。\n"
            "直前でも、まだ間に合います。\n\n"
            "商品と価格を見る（CTA）"
        ),
        "note": "※以前の投稿を参考に、ポップ目の印象が良さそうです",
        "url": f"{BASE}/lineup/",
        "campaign": "20260918_message_SchoolFestival",
    },
    {
        "no": 27,
        "date": "2026/09/29",
        "draft": "2026/09/28",
        "kind": "通常配信",
        "title": "新商品・トレカ台紙／推し活グッズ",
        "aim": (
            "Meta広告の主力フックである推し活を、LINE側でも正面から扱う唯一の全員配信。"
            "新商品告知は開封理由になり、BtoC層に「この会社は自分向けでもある」と"
            "認識させる役割を持つ（現状ここが弱くブロック要因になっている）。"
        ),
        "purpose": "CV獲得",
        "target": "BtoC・推し活層・TCGショップ・同人＋BtoBのノベルティ担当",
        "audience": "全員",
        "banner": (
            "既存のトレカ台紙バナーを流用可。4タイプ（A〜D）の並列構成はそのまま活かす。"
            "「1種類1枚から」を追加。下部CTA帯を赤に統一。"
        ),
        "text": (
            "推しのカード、\n"
            "そのまま渡すのはもったいない。\n\n"
            "トレーディングカード用の台紙、\n"
            "4タイプから選べます。\n\n"
            "1種類1枚から。\n"
            "推し活にも、ショップの販促にも。\n\n"
            "台紙を見る（CTA）"
        ),
        "note": "※既存バナー流用可",
        "url": f"{BASE}/lineup/trading_card_all/",
        "campaign": "20260929_message_TradingCardMount",
    },
    # --- 9月キャンペーン3本立て（内容確定後に日付と数字を埋める）---
    # 8月の型（開始告知 → 中盤リマインド → 最終日リマインド）を踏襲。
    # ◯ の箇所はキャンペーン名・期間・割引率が決まり次第そのまま置き換える。
    {
        "no": 28,
        "date": "要確定",
        "draft": "要確定",
        "kind": "通常配信",
        "title": "◯◯キャンペーン（開始告知）",
        "aim": (
            "期間限定の割引を告知して、期間全体の注文を立ち上げる。"
            "「データのご入稿は後から」を最初から明示し、初動の摩擦を消す。"
        ),
        "purpose": "キャンペーン訴求",
        "target": "BtoB・店舗・法人（対象商材が名刺・チラシ・冊子のため）",
        "audience": "全員",
        "banner": "※トンマナは後日確定。季節モチーフ背景／左上に黄丸で割引率／中央に対象商材／下部CTA帯／右下に期限",
        "text": (
            "◯◯キャンペーン、今日から始まります。\n\n"
            "最大◯%OFF。\n"
            "対象は名刺・ショップカード、\n"
            "チラシ・フライヤー、冊子です。\n\n"
            "◯月◯日まで。\n"
            "データのご入稿は後からで大丈夫です。\n"
            "先にご注文だけ進めておけます。\n\n"
            "対象商品を見る（CTA）"
        ),
        "note": "※キャンペーン名・期間・割引率が未確定",
        "url": f"{BASE}/cp/renewal/",
        "campaign": "2026MMDD_message_◯◯Campaign",
    },
    {
        "no": 29,
        "date": "要確定",
        "draft": "要確定",
        "kind": "通常配信",
        "title": "◯◯キャンペーン（中盤リマインド）",
        "aim": "「あとで」と先送りした層を動かす。残り日数を明示して締切を意識させる。",
        "purpose": "キャンペーン訴求",
        "target": "BtoB・店舗・法人",
        "audience": "全員",
        "banner": "※開始告知のバナーを流用し、左上を「残り◯日」に差し替え",
        "text": (
            "◯◯キャンペーン、残り◯日です。\n\n"
            "最大◯%OFF。\n"
            "名刺・チラシ・冊子が対象です。\n\n"
            "「あとで」と思っていた方、\n"
            "そろそろ動いておいてください。\n"
            "◯月◯日で終わります。\n\n"
            "対象商品を見る（CTA）"
        ),
        "note": "※開始告知バナーの流用で制作コストを抑える",
        "url": f"{BASE}/cp/renewal/",
        "campaign": "2026MMDD_message_◯◯CampaignRemind",
    },
    {
        "no": 30,
        "date": "要確定",
        "draft": "要確定",
        "kind": "通常配信",
        "title": "◯◯キャンペーン（最終日リマインド）",
        "aim": (
            "最終日の離脱理由の大半は「データが間に合わない」。"
            "注文と入稿を切り離せることを明示して、そこだけを潰す。"
        ),
        "purpose": "キャンペーン訴求",
        "target": "BtoB・店舗・法人",
        "audience": "全員",
        "banner": "※中盤バナーの「残り◯日」を「本日終了」に差し替え",
        "text": (
            "本日23:59で終了します。\n\n"
            "◯◯キャンペーン、最大◯%OFF。\n"
            "今日中のご注文までが対象です。\n\n"
            "データはまだ無くても大丈夫です。\n"
            "ご注文だけ先に済ませてください。\n\n"
            "今すぐ注文する（CTA）"
        ),
        "note": "※終了時刻（23:59か締切時間か）を要確認",
        "url": f"{BASE}/cp/renewal/",
        "campaign": "2026MMDD_message_◯◯CampaignLast",
    },
]


def final_url(url: str, campaign: str) -> str:
    return (
        f"{url}?utm_source={UTM_SOURCE}&utm_medium={UTM_MEDIUM}&utm_campaign={campaign}"
    )


def build() -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "9月配信企画案"

    header_fill = PatternFill("solid", fgColor=GREEN)
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28

    for r, p in enumerate(PLANS, start=2):
        row = [
            p["no"], p["date"], p["draft"], "草案作成中", p["kind"],
            p["title"], p["aim"], p["purpose"], p["target"], p["audience"],
            "1040×1040 px", p["banner"], p["text"], p["note"],
            p["url"], UTM_SOURCE, UTM_MEDIUM, p["campaign"],
            final_url(p["url"], p["campaign"]),
        ]
        for i, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.font = Font(size=10)
        # セグメント配信の行に色を敷いて、全員配信と見分けられるようにする
        if p["kind"] == "セグメント配信":
            for i in range(1, len(COLUMNS) + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GRAY)
        ws.row_dimensions[r].height = 150

    ws.freeze_panes = "B2"

    out = "プリントアース_LINEOA_配信企画案_2026-09.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    print("生成:", build())
