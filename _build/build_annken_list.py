# -*- coding: utf-8 -*-
"""佐村さん指定の11列で、案件一覧をそのまま貼れる形にする。

  python _build/build_annken_list.py <月次管理ブック.xlsx> [出力先ベース名]
  オプション:
    --アカウント貸し込み   既定では外している。外さずに全部入れたいとき
    --OA除外               LINE公式アカウントの媒体を外す

佐村さんの指示（2026-08-31 Talknote）
  「部署集計の各商材の
    商材　クライアント名　エンドクライアント名　商材詳細
    ストック売上　ストック利益　売上　利益　担当者1　担当者2　担当者3
    を合わせてもらうだけ」
  「関数とかいらないんだよね…」

  → 1行=1案件。集計もサマリも作らない。数式も入れない。値だけ。

列の作り方
  商材            月次管理ブックの `商材`
  クライアント名   `社名`（請求先）
  エンドクライアント名 `エンドクライアント名`（空なら社名で埋める）
  商材詳細         `サービス詳細`
  ストック売上     `計上種別` が「ストック」の案件だけ `請求額（税抜）`。他は空欄
  ストック利益     同上の `利益`
  売上            `請求額（税抜）`（契約形態に関係なく全部）
  利益            `利益`
  担当者1〜3       `アカウント` `コンサル` `運用①②③` の順に、重複を除いて3人まで
                  （ASPだけ列名が `広告主担当①②` `媒体担当①②` なのでそちらも見る）

出力
  <ベース名>.xlsx        見出し付き。中身の確認用
  <ベース名>.tsv         見出し無し。既にある表の途中に貼るとき用
  <ベース名>_見出し付き.tsv  1行目が見出し。新しいタブを作って貼るとき用

元ブックは読むだけで一切変更しない。
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_attack_list import MEDIA, extract

FONT = "メイリオ"
NAVY = "1F3864"
YEN = "#,##0;-#,##0;\"-\""
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# 佐村さん指定の並び。ここを変えると貼り付け位置がずれるので触らない
HEADERS = ["商材", "クライアント名", "エンドクライアント名", "商材詳細",
           "ストック売上", "ストック利益", "売上", "利益",
           "担当者1", "担当者2", "担当者3"]
WIDTHS = [18, 28, 28, 20, 14, 14, 14, 14, 14, 14, 14]
MONEY = (5, 6, 7, 8)

KAIGASHI = "アカウント貸し"
LINE_MEDIA = "LINE公式アカウント"


def f(sz=10, bold=False, color="000000"):
    return Font(name=FONT, size=sz, bold=bold, color=color)


def rows_from(recs):
    """1案件 → 11列の1行。"""
    out = []
    for r in recs:
        stock = r["契約形態"] == "ストック"
        tanto = [t for t in r["担当"].split(" / ") if t]
        out.append([
            r["商材カテゴリ"],
            r["社名"],
            r["エンドクライアント名"],
            r["サービス詳細"],
            r["売上"] if stock else None,      # ストック売上
            r["粗利"] if stock else None,      # ストック利益
            r["売上"],
            r["粗利"],
            tanto[0] if len(tanto) > 0 else "",
            tanto[1] if len(tanto) > 1 else "",
            tanto[2] if len(tanto) > 2 else "",
        ])
    return out


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "案件一覧"

    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = f(10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[1].height = 24

    for j, row in enumerate(rows):
        rr = 2 + j
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=rr, column=i, value=v)
            c.font = f(9)
            c.border = BORDER
            c.alignment = Alignment(vertical="top")
        for i in MONEY:
            ws.cell(row=rr, column=i).number_format = YEN

    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{1 + len(rows)}"
    wb.save(path)


def write_tsv(path, rows, header=False):
    """そのまま貼れるTSVを書く。header=True で1行目に見出しを付ける。

    元データのセルにタブや改行が入っていることがある
    （例：「ワンタグシステム利用料[タブ]_月額固定費」）。
    そのまま出すと列がずれるので、半角スペースに潰す。
    """
    with open(path, "w", encoding="utf-8") as fh:
        if header:
            fh.write("\t".join(HEADERS) + "\n")
        for row in rows:
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, float):
                    cells.append(f"{v:.0f}")
                else:
                    s = str(v)
                    for ch in ("\t", "\r\n", "\r", "\n"):
                        s = s.replace(ch, " ")
                    cells.append(s.strip())
            fh.write("\t".join(cells) + "\n")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    keep_kashi = "--アカウント貸し込み" in sys.argv
    drop_oa = "--OA除外" in sys.argv
    if not args:
        print(__doc__)
        sys.exit(1)
    src = args[0]
    base = args[1] if len(args) > 1 else "案件一覧_部署集計"

    recs, skipped = extract(src)
    print(f"読み込み        : {len(recs):,}件")

    if not keep_kashi:
        before = len(recs)
        recs = [r for r in recs if KAIGASHI not in r["サービス詳細"]]
        print(f"アカウント貸し除外: {before - len(recs)}件")
    if drop_oa:
        before = len(recs)
        recs = [r for r in recs if r["媒体"] != LINE_MEDIA]
        print(f"OA除外          : {before - len(recs)}件")

    # 媒体の並び順 → 同じ媒体の中は売上の大きい順
    order = {m: i for i, m in enumerate(MEDIA)}
    recs.sort(key=lambda r: (order.get(r["媒体"], 99), -r["売上"]))

    rows = rows_from(recs)
    write_xlsx(f"{base}.xlsx", rows)
    write_tsv(f"{base}.tsv", rows)
    write_tsv(f"{base}_見出し付き.tsv", rows, header=True)

    stock = [r for r in rows if r[4] is not None]
    print(f"書き出しました  : {base}.xlsx / {base}.tsv / {base}_見出し付き.tsv")
    print(f"  案件数        : {len(rows):,}")
    print(f"  売上          : {sum(r[6] for r in rows):,.0f}")
    print(f"  利益          : {sum(r[7] for r in rows):,.0f}")
    print(f"  うちストック  : {len(stock):,}件 "
          f"売上 {sum(r[4] for r in stock):,.0f} / 利益 {sum(r[5] for r in stock):,.0f}")
    print(f"  担当者1が空欄 : {sum(1 for r in rows if not r[8])}件")
    for sheet, miss in skipped:
        print(f"  ★スキップ {sheet}: {miss}")


if __name__ == "__main__":
    main()
