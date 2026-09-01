# -*- coding: utf-8 -*-
"""賃貸業界 Ver.B（オーナー開拓版）資料 公表判断チェックリスト（xlsx）

  python _build/build_checklist_verB.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説。データが入れば緑になる（公表可だが「未取得」明記が必要）
  赤 = 出典未記載の数値が載っている（出典を確定 or 削除するまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "賃貸業界資料_公表判断チェックリスト_VerB.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")
Y_FILL = PatternFill("solid", fgColor="FFF2CC")
R_FILL = PatternFill("solid", fgColor="FCE4E4")
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "検討背景", "事実", "緑", "LINEヤフー公式（国内MAU・開封率）。GMO調査の反証材料を明記し消去法で構成", "―"),
    (4, "管理受託のWebマーケ構造", "意見", "緑", "DYM提案の構造整理。数値なし", "―"),
    (5, "オーナーの心理", "意見", "緑", "DYM仮説。数値なし", "上司確認で裏付けを取ること"),
    (6, "できること／できないこと＋スコープ宣言", "意見", "緑", "DYM提案のスコープ宣言。数値なし", "―"),
    (7, "実態調査｜モーメント分析", "事実", "緑", "Googleトレンド実データ（相続5年・確定申告5年）", "提案時に取得日を更新"),
    (8, "実態調査｜前後検索", "事実", "緑", "LINEヤフー媒体資料（取得済み・反映済み）", "―"),
    (9, "他社分析①｜競合のLINE運用ステータス", "事実", "黄", "6社選定は事実。友だち数・オーナー向け別アカウントは★未取得", "page.line.meから取得日つきで実測"),
    (10, "他社分析②｜運用の型4分類", "推定", "黄", "4分類は一般的な運用パターン。各社の実際の型は★未実測", "実機スクリーンショットで型を確認"),
    (11, "全体設計｜カスタマージャーニー", "意見", "緑", "DYM標準フレームにオーナー開拓を当てはめ", "―"),
    (12, "全体設計｜施策全体像", "意見", "緑", "DYM提案の施策全体像", "―"),
    (13, "全体設計｜効果を最大化する2軸", "意見", "緑", "新規・既存の2軸設計。数値なし", "―"),
    (14, "全体設計｜対策領域マップ", "意見", "緑", "新規開拓・既存維持の2領域", "―"),
    (15, "全体設計｜施策展開図（初期・月次）", "意見", "緑", "DYM標準の展開図をオーナー開拓施策に置換", "―"),
    (16, "構築｜友だち追加動線", "意見", "緑", "DYM提案の動線設計", "―"),
    (17, "構築｜あいさつメッセージ", "意見", "緑", "配信文面は初稿", "担当者名・実際の運用フローに差し替え"),
    (18, "構築｜土地活用タイプ診断", "意見", "緑", "診断設計。S08前後検索データに基づき順番を設計", "―"),
    (19, "構築｜リッチメニュー", "意見", "緑", "DYM提案のリッチメニュー設計", "―"),
    (20, "構築｜長期育成の設計思想", "意見", "緑", "DYM提案の設計思想。数値なし", "―"),
    (21, "配信設計｜年次モーメントカレンダー", "事実", "緑", "確定申告のピークはGoogleトレンド実データ（S07：3→100・33倍）", "―"),
    (22, "配信設計｜企画投稿案（月次）", "意見", "緑", "DYM提案の企画投稿設計", "―"),
    (23, "配信設計｜初動30日のステップ", "意見", "緑", "配信設計。数値なし", "―"),
    (24, "配信設計｜実文面", "意見", "緑", "配信文面は初稿", "実際の事例・診断結果に差し替え"),
    (25, "配信設計｜セミナー送客", "意見", "緑", "DYM提案のセミナー送客設計", "―"),
    (26, "配信設計｜通知メッセージ", "意見", "緑", "利用シーンの設計", "―"),
    (27, "歩留まり・工数｜相談の歩留まり改善", "意見", "緑", "DYM提案の歩留まり改善設計", "―"),
    (28, "歩留まり・工数｜工数削減", "意見", "緑", "DYM提案の工数削減設計", "―"),
    (29, "歩留まり・工数｜改善モデル", "意見", "緑", "DYM提案の改善モデル", "―"),
    (30, "成果と体制｜効果測定の設計", "意見", "緑", "数値は入れない（業界汎用のためSIMは作らない）。主KPI＝面談CPA・CVR", "御社の実績をいただければSIMを作成"),
    (31, "成果と体制｜費用プラン", "意見", "緑", "DYM標準プラン（6ヶ月〜・税抜）", "―"),
    (32, "成果と体制｜運用スケジュール", "意見", "緑", "DYM標準の運用スケジュール", "―"),
    (33, "成果と体制｜サポート体制", "意見", "緑", "DYM標準のサポート体制", "―"),
    (34, "締め｜飛び道具", "事実", "緑", "A・D案は前後検索データ（S08）で裏付け済み。B・Cは意見（施策アイデア）", "―"),
    (35, "締め｜LINEOA実績", "事実", "緑", "オーナー向け公式事例は0件（2回探索済み・正直に明記）。近接事例はLINEヤフー公式", "掲載前に各事例ページの原典で再確認"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（原本のまま）", "―"),
]
assert len(ROWS) == 36, len(ROWS)

wb = Workbook()
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "賃貸業界_LINEOA施策提案_VerB_オーナー開拓（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-31　対象ファイル：賃貸業界_LINEOA施策提案_VerB_オーナー開拓.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("オーナーは営業されたくないが、情報は欲しい。しかも検討は1〜3年。"
            "訪問・電話・DMは「その瞬間に検討中の人」しか拾えない。"
            "やる意思はあるが今すぐではない層を育て続ければ、面談獲得の単価（CPA）が下がる。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（未取得明記のうえ公表可）", "黄", Y_FILL),
    ("赤（出典確定まで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = '=IF(D11=0,"OK：赤0件。対外提出可（黄は未取得データの明記つきで提出可）","NG：赤"&D11&"件。出典を確定または該当数値を削除するまで対外提出不可")'
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 上司確認が必要な3点"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① オーナーの心理（S05）の裏付け確認（DYM仮説として置いた。実案件のヒアリング結果があれば差し替え）",
    "② GMO調査（オーナーはメール首位）の一次ソース再確認・S03での反証材料の扱い方針",
    "③ 競合6社の実名掲載OK（S09）→ NGなら「A社／B社」表記に変更",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B22"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. Googleトレンド（取得済み。相続・確定申告とも実データ反映済み。S07）",
    "2. 上司確認3点セット（上記★）→ S05・S09が黄→緑",
    "3. LINEヤフーの前後検索データ（取得済み・反映済み。S08）",
    "4. 競合の友だち数実測（大東建託／東建コーポレーション／生和コーポレーション／シノケン／積水ハウス不動産／大和ハウス。オーナー向け別アカウントの有無も記録）→ S09・S10が黄→緑",
    "5. 事例ストック調査結果／上司共有資料の反映（オーナー向け公式事例がないため代替材料があれば強化できる）→ S35の裏取り",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B29"] = "凡例（このファイルの使い方）"
ws["B29"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
]):
    ws.cell(30 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 30

for col, w in zip("ABCDEF", [5, 42, 11, 7, 52, 46]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
