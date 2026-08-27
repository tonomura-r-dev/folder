# -*- coding: utf-8 -*-
"""注文住宅業界資料 公表判断チェックリスト（xlsx）

  python _build/build_checklist.py
  python /root/.claude/skills/synced/xlsx/scripts/recalc.py 注文住宅業界資料_公表判断チェックリスト.xlsx

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説。データが入れば緑になる（公表可だが「仮説」明記が必要）
  赤 = 出典未記載の数値が載っている（出典を確定 or 削除するまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DECK = ROOT / "注文住宅業界_LINEOA施策提案.pptx"
OUT = ROOT / "注文住宅業界資料_公表判断チェックリスト.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑
Y_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # 赤
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")  # 記入セル
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (枚, タイトル, 主張の種類, 判定, 根拠・出典, 残タスク)
ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "LINE公式アカウントを検討する背景", "意見", "緑", "DYM標準の課題整理（FMT）", "―"),
    (3, "全体設計想定（カスタマージャーニー）", "意見", "緑", "CV3段はDYM設計（本提案で定義）", "貴社のCV定義とすり合わせ"),
    (4, "全体設計想定（簡略版）", "意見", "緑", "同上", "―"),
    (5, "市場環境｜検索広告CPCの推移", "推定", "黄", "CPCは業界水準に基づく推計レンジ（脚注に明記）", "貴社の実績CPCに差し替え（上司確認①）"),
    (6, "広告与件｜CV地点別の設計マトリクス＋広告審査", "推定", "黄", "想定CVRは業界水準。CV地点は各行に明記", "貴社の実績CVR/CPAに差し替え（上司確認①）"),
    (7, "ニーズ調査（シーズナリティ）", "推定", "黄", "指数は業界水準の推計（年平均=100・脚注に明記）", "Googleトレンド実データ（5年＋1年）で差し替え"),
    (8, "ニーズ調査（前後検索）", "推定", "黄", "業界知見に基づく仮説（脚注に明記）", "LINEヤフーの前後検索データで検証"),
    (9, "他社分析①｜主要8社のLINE運用ステータス", "事実", "黄", "アカウント実在は確認済。友だち数は未取得＝差込枠のまま", "page.line.meで取得日つき実測（DYM対応）"),
    (10, "他社分析②｜運用の型4分類", "意見", "黄", "型の分類はDYM見解。仮説バッジを表示済", "実測後に各社をプロットして確定"),
    (11, "資料アジェンダ", "意見", "緑", "構成の目次", "―"),
    (12, "具体企画①｜ステップ配信案（14日）", "意見", "緑", "配信設計。根拠列は前後検索（S8）に紐づく", "S8確定後に根拠列を実データ表記へ"),
    (13, "具体企画②｜実際にスマホへ届く文面", "意見", "緑", "配信文面の初稿。数値は登場しない", "担当者名・実績値を貴社データに差し替え"),
    (14, "具体企画③｜LINE通知メッセージの利用シーン", "意見", "緑", "利用シーンの設計", "電話番号の取得同意状況を確認"),
    (15, "具体企画④｜飛び道具", "意見", "緑", "企画案。数値なし", "B案（不利の開示）は社内合意が必要"),
    (16, "LINEOA実績｜LINEヤフー公式の導入事例", "事実", "黄", "LINEヤフー公式 導入事例（lycbiz.com）記載の数値", "掲載前に各事例ページ原典で数値を再確認"),
    (17, "LINE施策における重要な考え方", "意見", "緑", "DYM標準の考え方（FMT）", "―"),
    (18, "LINE公式アカウントを検討する背景（集客効率）", "推定", "赤", "「業態平均で最大3%」「3%/1.5%/2.5%/2%」＝出典未記載", "出典を社内で確定 or 該当数値を削除（上司確認②）"),
    (19, "LINE公式アカウントにおける対策領域", "意見", "緑", "機能マップ（FMT）", "―"),
    (20, "施策展開図（初期・月次）", "意見", "緑", "実施メニュー（FMT）", "―"),
    (21, "施策展開図（友だち追加後の動線）", "意見", "緑", "8ブロックを注文住宅の施策に置換済", "―"),
    (22, "施策展開図（15日でアクティブ率が下がる）", "推定", "黄", "「15日を境にアクティブ率が下がる」＝DYM運用知見", "出典ラベルを付す（社内実績 or 業界水準）"),
    (23, "施策の考え方（成果地点を軸に配信）", "意見", "緑", "配信戦略の考え方（FMT）", "―"),
    (24, "改善モデル①（個人情報の最大化）", "推定", "赤", "「CVR 2%前後を5〜10%に引上げ」＝出典未記載", "出典を社内で確定 or 該当数値を削除（上司確認②）"),
    (25, "改善モデル②（リードの直リーチ化）", "推定", "赤", "「商談化70%改善事例も」「再CV率3〜8%」＝出典未記載", "出典を社内で確定 or 該当数値を削除（上司確認②）"),
    (26, "改善モデル③（マネタイズ）", "意見", "緑", "施策設計。注文住宅の文脈に置換済", "―"),
    (27, "改善モデル④（LTV最大化）", "意見", "緑", "外構・点検・紹介の文脈に置換済。出典不明数値は削除済", "―"),
    (28, "資料アジェンダ", "意見", "緑", "構成の目次", "―"),
    (29, "具体施策（低ハードルCV地点の設定）", "推定", "赤", "「友だち追加率を0.8〜1.0%前後に改善」＝出典未記載", "出典を社内で確定 or 該当数値を削除（上司確認②）"),
    (30, "具体施策（保有リストの再CV）", "意見", "緑", "施策設計。注文住宅の文脈に置換済", "―"),
    (31, "具体施策（紹介・追加工事の計画実施）", "意見", "緑", "同上。出典不明数値は削除済", "―"),
    (32, "具体施策（チャット簡略化・営業トスアップ）", "意見", "緑", "施策設計（FMT）", "―"),
    (33, "具体施策（短期）｜60秒 予算診断の設問設計", "意見", "緑", "設問設計。数値なし", "―"),
    (34, "具体施策（長期）｜年間サイクルの企画投稿案", "推定", "黄", "時期設定は業界水準の年間サイクルに基づく推計", "Googleトレンド実データでピーク月を調整"),
    (35, "具体施策（長期）｜診断タイプ別の追客強弱", "意見", "緑", "配信設計。数値なし", "運用2ヶ月目のブロック率実測で頻度調整"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報", "―"),
]

wb = Workbook()

# ---------------- サマリ ----------------
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "注文住宅業界_LINEOA施策提案（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-26　対象ファイル：注文住宅業界_LINEOA施策提案.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("注文住宅は検討期間が8〜12ヶ月あるのに、御社が接点を持てるのは「フォームを通れた人」だけ。"
            "LPから落ちた後と予約が入った後の2箇所だけをLINEで埋め、来場率と契約棟数を引き上げる。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（仮説明記のうえ公表可）", "黄", Y_FILL),
    ("赤（出典確定まで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = '=IF(D11=0,"OK：赤0件。対外提出可","NG：赤"&D11&"件。出典を確定または該当数値を削除するまで対外提出不可")'
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 上司確認が必要な3点"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① 社内広告実績の共有可否（CV地点別のCPC/CVR/CPA・資料請求→来場率・来場→契約率・1棟粗利率）→ S5/S6が実測になる",
    "② 出典未確定数値の扱い（S18・S24・S25・S29の赤4枚）→ 出典を確定すれば緑、できなければ該当数値を削除",
    "③ 競合8社・導入事例の実名掲載OK（S9/S10/S16）→ NGなら「A社／B社」表記に変更",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B22"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. Googleトレンドのスクショ（注文住宅／ハウスメーカー／注文住宅 相場／土地 探し方・5年＋直近1年）→ S7・S34が黄→緑",
    "2. 上司確認3点セット（上記★）→ S5・S6・S18・S24・S25・S29が黄／赤→緑",
    "3. LINEヤフーの前後検索データ（注文住宅 相場／ハウスメーカー 比較／注文住宅 ●●市）→ S8が黄→緑",
    "4. 競合の友だち追加調査（上位3社のあいさつ・リッチメニュー・2週間の配信内容）→ S10が黄→緑",
    "5. 事例ストック調査結果／上司共有資料の反映 → S16の裏取り",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B29"] = "凡例（このファイルの使い方）"
ws["B29"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
]):
    ws.cell(30 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

# ---------------- 判定一覧 ----------------
ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 30

for col, w in zip("ABCDEF", [5, 42, 11, 7, 52, 46]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
