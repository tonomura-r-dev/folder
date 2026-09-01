# -*- coding: utf-8 -*-
"""EC業界資料 公表判断チェックリスト（xlsx）

  python _build/build_checklist_ec.py

全36枚を「事実／推定／意見」で仕分け、赤黄緑で公表可否を判定する。
判定基準：
  緑 = 出典のある事実、またはDYMの標準メニュー・設計（そのまま公表可）
  黄 = 推定・仮説。データが入れば緑になる（公表可だが「仮説」「未取得」明記が必要）
  赤 = 出典未記載の数値が載っている（出典を確定 or 削除するまで公表不可）
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
DECK = ROOT / "EC業界_LINEOA施策提案.pptx"
OUT = ROOT / "EC業界資料_公表判断チェックリスト.xlsx"

FONT = "Arial"
NAVY = "1F285A"
G_FILL = PatternFill("solid", fgColor="E2EFDA")   # 緑
Y_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄
R_FILL = PatternFill("solid", fgColor="FCE4E4")   # 赤
HDR_FILL = PatternFill("solid", fgColor=NAVY)
IN_FILL = PatternFill("solid", fgColor="FFFF00")  # 記入セル
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (枚, タイトル, 主張の種類, 判定, 根拠・出典, 残タスク)
ROWS = [
    (1, "表紙", "意見", "緑", "提案タイトル。数値なし", "宛先社名を差し込む"),
    (2, "資料アジェンダ", "意見", "緑", "全8章の目次", "―"),
    (3, "市場データ｜ニーズ全体図", "事実", "緑", "経産省「令和6年度 電子商取引に関する市場調査」2025年8月26日公表", "提案時に最新年度（令和7年度）を確認"),
    (4, "ECのWebマーケ構造｜3つの壁", "事実", "緑", "カゴ落ち率：Baymard Institute／イー・エージェンシー。CPCはキーワードマーケティング調べ", "―"),
    (5, "EC購入者の心理", "事実", "緑", "カゴ落ち理由：Baymard Institute。メルマガ開封率は業界一般値", "―"),
    (6, "できること／できないこと＋スコープ宣言", "意見", "緑", "DYM提案のスコープ宣言。数値なし", "―"),
    (7, "市場環境｜検索広告CPCの推移", "事実", "緑", "電通「2025年 日本の広告費」2026年3月5日発表／キーワードマーケティング", "―"),
    (8, "市場環境｜広告実績CVR（CV地点記載）", "推定", "黄", "★差込枠。数値は未記入（捏造なし）", "DYM社内のEC広告実績をCV地点別に取得（上司確認①）"),
    (9, "市場環境｜広告審査面の懸念", "事実", "緑", "薬機法の広告3要件（薬監発第148号）／京都府「広告の３要件」", "薬機法の審査落ち実例（社内1〜2例）があれば説得力が上がる（任意）"),
    (10, "ニーズ調査｜シーズナリティ", "推定", "黄", "年間商戦カレンダーは公開の商戦・セール時期（事実）。Googleトレンドは★差込枠", "Googleトレンド実データ（福袋／ブラックフライデー／セール／母の日・一語・5年＋1年）"),
    (11, "ニーズ調査｜前後検索", "推定", "黄", "★差込枠。KWのみ提示、データ未記入", "LINEヤフー前後検索データ（通販／定期便／解約／口コミ）"),
    (12, "他社分析｜競合10社のLINE運用ステータス", "事実", "黄", "10社選定は知名度優先（事実）。友だち数・ID連携は★未取得", "page.line.meから取得日つきで実測（友だち数・ID連携）"),
    (13, "全体設計｜カスタマージャーニー×CV3段", "意見", "緑", "CV3段はDYM設計。数値なし", "貴社のCV定義とすり合わせ"),
    (14, "全体設計｜施策全体像・対策領域マップ", "意見", "緑", "DYM提案の対策領域マップ", "―"),
    (15, "全体設計｜施策展開図（初期・月次）", "意見", "緑", "DYM標準の展開図をEC施策に置換", "―"),
    (16, "構築｜友だち追加動線", "意見", "緑", "DYM提案の動線設計", "―"),
    (17, "構築｜あいさつメッセージ", "意見", "緑", "配信文面は初稿。カゴ落ち理由の出典はS5", "担当者名・実際のクーポン条件に差し替え"),
    (18, "構築｜診断・アンケート", "意見", "緑", "診断設計。数値なし", "―"),
    (19, "構築｜リッチメニュー", "意見", "緑", "DYM提案のリッチメニュー設計", "―"),
    (20, "構築｜LINE ID連携（★心臓部）", "意見", "緑", "仕組みの説明。数値なし", "―"),
    (21, "配信設計｜シナリオ2本の設計表", "意見", "緑", "配信設計。日数は初稿（★想定）", "F2オファーのタイミングは商材ごとに調整"),
    (22, "配信設計｜実文面①｜未購入者", "意見", "緑", "配信文面は初稿", "実際のクーポン条件・期限に差し替え"),
    (23, "配信設計｜実文面②｜購入者", "意見", "緑", "配信文面は初稿", "―"),
    (24, "配信設計｜カゴ落ちリマインドの設計", "意見", "緑", "通数・タイミングは初稿（★想定）", "実運用データで配信間隔を調整"),
    (25, "配信設計｜年間の企画投稿カレンダー", "推定", "黄", "商戦カレンダー（S10）に基づく企画案", "S10のGoogleトレンド確定後に時期を微調整"),
    (26, "配信設計｜通知メッセージ", "意見", "緑", "利用シーンの設計", "電話番号・LINE ID取得の同意状況を確認"),
    (27, "歩留まり・工数｜F2転換率の改善（★本資料の山）", "事実", "緑", "F2転換率・定期引上率は業界水準（出典明記）", "―"),
    (28, "歩留まり・工数｜工数削減＋改善モデル", "意見", "緑", "DYM提案の改善モデル", "―"),
    (29, "成果と体制｜効果測定の設計", "意見", "緑", "数値は入れない（業界汎用のためSIMは作らない）", "御社の実績をいただければSIMを作成"),
    (30, "成果と体制｜費用プラン", "意見", "緑", "DYM標準プラン（6ヶ月〜・税抜）", "―"),
    (31, "成果と体制｜運用スケジュール", "意見", "緑", "DYM標準の運用スケジュール", "―"),
    (32, "成果と体制｜サポート体制", "意見", "緑", "DYM標準のサポート体制", "―"),
    (33, "締め｜飛び道具", "意見", "黄", "A〜Cは施策アイデア（意見）。D案は仮説バッジ付き", "D案はS11（前後検索）で「解約」不安の実在を確認後に本提案化"),
    (34, "締め｜第2の提案軸｜LINEミニアプリ・LINEギフト", "意見", "緑", "DYM提案の第2軸", "―"),
    (35, "締め｜LINEOA実績", "事実", "黄", "LINE国内MAU・開封率はLINEヤフー公式値（緑相当）。EC事例は★未取得", "lycbiz.com EC専用ページから事例取得"),
    (36, "裏表紙", "事実", "緑", "DYM会社情報（原本のまま）", "―"),
]
assert len(ROWS) == 36, len(ROWS)

wb = Workbook()

# ---------------- サマリ ----------------
ws = wb.active
ws.title = "サマリ"
ws.sheet_view.showGridLines = False

ws["B2"] = "EC業界_LINEOA施策提案（36枚）｜公表判断チェックリスト"
ws["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws["B3"] = "作成日：2026-08-31　対象ファイル：EC業界_LINEOA施策提案.pptx"
ws["B3"].font = Font(name=FONT, size=9, color="7F7F7F")

ws["B5"] = "背骨（この資料が言っていること・1文）"
ws["B5"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B6"] = ("新規獲得の単価は上がり続けている。もう「集める」では勝てない。Webマーケでできるのは、"
            "いま取りこぼしている購入と、一度きりで終わっている顧客を、追加の広告費ゼロで拾い直すこと。"
            "勝負はCPOではなく、LTV÷CPO。")
ws["B6"].font = Font(name=FONT, size=10)
ws["B6"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B6:F6")
ws.row_dimensions[6].height = 32

ws["B8"] = "判定サマリ"
ws["B8"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, (lab, cond, fill) in enumerate([
    ("緑（そのまま公表可）", "緑", G_FILL),
    ("黄（未取得・仮説明記のうえ公表可）", "黄", Y_FILL),
    ("赤（出典確定まで公表不可）", "赤", R_FILL),
]):
    r = 9 + i
    ws.cell(r, 2, lab).font = Font(name=FONT, size=10)
    c = ws.cell(r, 4, f'=COUNTIF(判定一覧!$D$3:$D$38,"{cond}")')
    c.font = Font(name=FONT, size=10, bold=True)
    c.alignment = Alignment(horizontal="center")
    c.fill = fill
    c.border = BOX
    ws.cell(r, 5, "枚").font = Font(name=FONT, size=10)
ws.cell(12, 2, "合計").font = Font(name=FONT, size=10, bold=True)
c = ws.cell(12, 4, "=SUM(D9:D11)")
c.font = Font(name=FONT, size=10, bold=True)
c.alignment = Alignment(horizontal="center")
c.border = BOX
ws.cell(12, 5, "枚").font = Font(name=FONT, size=10)

ws["B14"] = "納品条件"
ws["B14"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
ws["B15"] = '=IF(D11=0,"OK：赤0件。対外提出可（黄は未取得データの明記つきで提出可）","NG：赤"&D11&"件。出典を確定または該当数値を削除するまで対外提出不可")'
ws["B15"].font = Font(name=FONT, size=10, bold=True, color="C00000")
ws.merge_cells("B15:F15")

ws["B17"] = "★ 上司確認が必要な3点"
ws["B17"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "① DYM社内のEC広告実績（カート投入／初回購入／定期引上のCV地点別CPC・CVR・CPO）→ S8が実測になる",
    "② 出典未確定・仮説の扱い（S33のD案「解約」不安）→ S11のデータで裏付けが取れれば本提案化、取れなければ仮説のまま or 削除",
    "③ 競合10社・LY公式EC事例の実名掲載OK（S12／S35）→ NGなら「A社／B社」表記に変更",
]):
    r = 18 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B22"] = "△ データ待ち（人が集める素材・5種に限定）"
ws["B22"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "1. Googleトレンド（福袋／ブラックフライデー／セール／母の日・一語・5年＋1年。「セール」は単独グラフ）→ S10が黄→緑",
    "2. 上司確認3点セット（上記★）→ S8・S12・S33・S35が黄→緑",
    "3. LINEヤフーの前後検索データ（通販／定期便／解約／口コミ）→ S11が黄→緑",
    "4. 競合の友だち数実測（10社。page.line.meから取得日つき。ID連携の有無も記録）→ S12が黄→緑",
    "5. LINEヤフー公式のEC導入事例（lycbiz.com/jp/service/line-official-account/case-study/ec/）→ S35が黄→緑",
]):
    r = 23 + i
    ws.cell(r, 2, t).font = Font(name=FONT, size=9.5)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

ws["B29"] = "凡例（このファイルの使い方）"
ws["B29"].font = Font(name=FONT, size=10, bold=True, color=NAVY)
for i, t in enumerate([
    "・「判定一覧」タブのF列（残タスク）が、この資料を完成させるためのTODOです。",
    "・黄色で塗られたセルが記入欄です。データが入ったらD列の判定を「緑」に書き換えてください。",
    "・D列を書き換えると、このサマリの集計と納品条件の判定が自動で更新されます。",
]):
    ws.cell(30 + i, 2, t).font = Font(name=FONT, size=9.5)

for col, w in zip("ABCDEF", [2.5, 62, 14, 10, 8, 40]):
    ws.column_dimensions[col].width = w

# ---------------- 判定一覧 ----------------
ws2 = wb.create_sheet("判定一覧")
ws2.sheet_view.showGridLines = False
hdr = ["枚", "スライドタイトル", "主張の種類", "判定", "根拠・出典", "残タスク（これが埋まれば緑）"]
for j, h in enumerate(hdr, start=1):
    c = ws2.cell(2, j, h)
    c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
ws2.freeze_panes = "A3"

FILLS = {"緑": G_FILL, "黄": Y_FILL, "赤": R_FILL}
for i, row in enumerate(ROWS):
    r = 3 + i
    for j, v in enumerate(row, start=1):
        c = ws2.cell(r, j, v)
        c.font = Font(name=FONT, size=9.5, bold=(j == 4))
        c.border = BOX
        c.alignment = Alignment(
            wrap_text=(j in (2, 5, 6)),
            horizontal="center" if j in (1, 3, 4) else "left",
            vertical="center")
    ws2.cell(r, 4).fill = FILLS[row[3]]
    if row[3] != "緑":
        ws2.cell(r, 6).fill = IN_FILL
    ws2.row_dimensions[r].height = 30

for col, w in zip("ABCDEF", [5, 42, 11, 7, 52, 46]):
    ws2.column_dimensions[col].width = w

wb.save(OUT)
print("saved:", OUT)
print("赤:", sum(1 for r in ROWS if r[3] == "赤"),
      "／黄:", sum(1 for r in ROWS if r[3] == "黄"),
      "／緑:", sum(1 for r in ROWS if r[3] == "緑"))
